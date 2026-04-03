import csv, os, zipfile, io,  tempfile
import pandas as pd
from io import StringIO
from io import BytesIO
from django.core.files import File
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import timedelta
from itertools import chain
from operator import attrgetter
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse
from django.contrib import messages
from django.db import IntegrityError
from django.contrib.auth.hashers import make_password, check_password

from .models import Admin, MayorsPermit, IDCard, Mtop, Franchise, MayorsPermitTricycle, SuperAdmin,MayorsPermitHistory, MayorsPermitTricycleHistory,ActivityLog,Tricycle, TricycleHistory
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from .forms import CSVUploadForm
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
import json
from django.views.decorators.http import require_http_methods, require_POST
from functools import wraps
import logging
from django.utils.timezone import now
from django.conf import settings
from django.db import transaction
import shutil
import openpyxl
from django.core.paginator import Paginator
from django.db.models import Q
from django.core.serializers import serialize
from django.utils.timezone import localtime

def superadmin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.session.get('user_type') != 'superadmin':
            messages.error(request, 'Access denied. Only Super Admins can access this page.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper

    

@require_POST
def add_mtop(request):

    case_no = request.POST.get("case_no")
    motor_no = request.POST.get("motor_no")
    chasses_no = request.POST.get("chasses_no")

    # ✅ FIELD-LEVEL VALIDATION
    if Mtop.objects.filter(case_no=case_no).exists():
        messages.error(request, f"Case No '{case_no}' already exists.")
        return JsonResponse({"success": False})

    if Mtop.objects.filter(motor_no=motor_no).exists():
        messages.error(request, f"Motor No '{motor_no}' already exists.")
        return JsonResponse({"success": False})

    if Mtop.objects.filter(chasses_no=chasses_no).exists():
        messages.error(request, f"Chassis No '{chasses_no}' already exists.")
        return JsonResponse({"success": False})

    try:
        mtop = Mtop.objects.create(
            name=request.POST.get("name"),
            case_no=case_no,
            address=request.POST.get("address"),
            no_of_units=request.POST.get("no_of_units"),
            route_operation=request.POST.get("route_operation"),
            make=request.POST.get("make"),
            motor_no=motor_no,
            chasses_no=chasses_no,
            plate_no=request.POST.get("plate_no"),
            date=request.POST.get("date"),
            municipal_treasurer=request.POST.get("municipal_treasurer"),
            officer_in_charge=request.POST.get("officer_in_charge"),
            mayor=request.POST.get("mayor"),
        )

        messages.success(request, f"{mtop.name} successfully added.")
        return JsonResponse({"success": True})

    except Exception as e:
        messages.error(request, "An unexpected error occurred while saving the record.")
        return JsonResponse({"success": False})


def franchise(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    franchises = Franchise.objects.all().order_by('-date')  # latest first
    return render(request, 'myapp/franchise.html', {'franchises': franchises}) 

def franchise_datatable(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    """Server-side processing endpoint for Franchise DataTables"""
    
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    
    queryset = Franchise.objects.select_related('tricycle').all().order_by('-date')

    if search_value:
        queryset = queryset.filter(
            Q(name__icontains=search_value) |
            Q(denomination__icontains=search_value) |
            Q(plate_no__icontains=search_value) |
            Q(motor_no__icontains=search_value) |
            Q(authorized_no__icontains=search_value) |
            Q(chassis_no__icontains=search_value) |
            Q(authorized_route__icontains=search_value) |
            Q(purpose__icontains=search_value) |
            Q(official_receipt_no__icontains=search_value) |
            Q(municipal_treasurer__icontains=search_value) |
            Q(valid_until__icontains=search_value) |
            Q(date__icontains=search_value) |
            Q(amount_paid__icontains=search_value) |
            Q(tricycle__body_number__icontains=search_value)  # ← ADDED
        )

    for i in range(16):  # ← was 14
        column_search = request.GET.get(f'columns[{i}][search][value]', '')
        if column_search:
            if i == 0:    queryset = queryset.filter(name__icontains=column_search)
            elif i == 1:  queryset = queryset.filter(denomination__icontains=column_search)
            elif i == 2:  queryset = queryset.filter(plate_no__icontains=column_search)
            elif i == 3:  queryset = queryset.filter(authorized_route__icontains=column_search)
            elif i == 4:  queryset = queryset.filter(motor_no__icontains=column_search)
            elif i == 5:  queryset = queryset.filter(authorized_no__icontains=column_search)
            elif i == 6:  queryset = queryset.filter(chassis_no__icontains=column_search)
            elif i == 7:  queryset = queryset.filter(tricycle__body_number__icontains=column_search)  # ← NEW
            elif i == 8:  queryset = queryset.filter(authorized_route__icontains=column_search)
            elif i == 9:  queryset = queryset.filter(purpose__icontains=column_search)
            elif i == 10: queryset = queryset.filter(official_receipt_no__icontains=column_search)
            elif i == 11: queryset = queryset.filter(date__icontains=column_search)
            elif i == 12: queryset = queryset.filter(amount_paid__icontains=column_search)
            elif i == 13: queryset = queryset.filter(municipal_treasurer__icontains=column_search)
    
    total_records = Franchise.objects.count()
    filtered_records = queryset.count()
    
    franchises = queryset[start:start + length]
    
    data = []
    for f in franchises:
        action_html = f'''
        <div class="btn-group" role="group" aria-label="actions">
            <button class="btn btn-sm btn-primary print-btn" 
                    data-permit-id="{f.id}"
                    data-name="{f.name}"
                    data-denomination="{f.denomination}"
                    data-plate="{f.plate_no}"
                    data-valid="{f.valid_until.strftime('%b-') + str(f.valid_until.day) + f.valid_until.strftime('-%Y')}"
                    data-motor="{f.motor_no}"
                    data-authorized="{f.authorized_no}"
                    data-chassis="{f.chassis_no}"
                    data-route="{f.authorized_route}"
                    data-purpose="{f.purpose}"
                    data-receipt="{f.official_receipt_no}"
                    data-date="{f.date.strftime('%b-') + str(f.date.day) + f.date.strftime('-%Y')}"
                    data-amount="₱{float(f.amount_paid):,.2f}"
                    data-treasurer="{f.municipal_treasurer}"
                    title="Print">
                <i class="fas fa-print"></i>
            </button>
            <a href="#" class="btn btn-sm btn-warning btn-franchise-update" data-id="{f.id}" data-body-number="{f.tricycle.body_number if f.tricycle else ''}" data-name="{f.name}" title="Update">
                <i class="fas fa-edit"></i>
            </a>
             <button class="btn btn-sm btn-danger btn-delete-franchise" title="Delete"
                data-id="{f.id}"
                data-name="{f.name}"
                data-plate="{f.plate_no}">
                <i class="fas fa-trash"></i>
            </button>
        </div>
        '''
        
        data.append([
            f.name,                                     # 0  - Name
            f.denomination,                             # 1  - Denomination
            f.plate_no,                                 # 2  - Plate No
            f.valid_until.strftime('%b-') + str(f.valid_until.day) + f.valid_until.strftime('-%Y'),  # 3  - Valid Until
            f.motor_no,                                 # 4  - Motor No
            f.authorized_no,                            # 5  - Authorized No
            f.chassis_no,                               # 6  - Chassis No
            f.tricycle.body_number if f.tricycle else '',  # 7  - Body Number ← NEW
            f.authorized_route,                         # 8  - Authorized Route (hidden)
            f.purpose,                                  # 9  - Purpose (hidden)
            f.official_receipt_no,                      # 10 - Official Receipt No (hidden)
            f.date.strftime('%b-') + str(f.date.day) + f.date.strftime('-%Y'),  # 11 - Date (hidden)
            '{:,.2f}'.format(float(f.amount_paid)),     # 12 - Amount Paid (hidden)
            f.municipal_treasurer,                      # 13 - Municipal Treasurer (hidden)
            action_html,                                # 14 - Action
            f.id,                                       # 15 - Hidden ID
            f.status,                                   # 16 - Status (hidden)
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })
    
# New view to add franchise
@require_http_methods(["POST"])
def add_franchise(request):
    try:
        # Get data from POST request
        name = request.POST.get('name')
        denomination = request.POST.get('denomination')
        plate_no = request.POST.get('plate_no')
        valid_until = request.POST.get('valid_until')
        motor_no = request.POST.get('motor_no')
        authorized_no = request.POST.get('authorized_no')
        chassis_no = request.POST.get('chassis_no')
        authorized_route = request.POST.get('authorized_route')
        purpose = request.POST.get('purpose')
        official_receipt_no = request.POST.get('official_receipt_no')
        date = request.POST.get('date')
        amount_paid = request.POST.get('amount_paid')
        municipal_treasurer = request.POST.get('municipal_treasurer')
        status = request.POST.get('status', 'New') 

        # Check for duplicate Plate Number
        if Franchise.objects.filter(plate_no=plate_no).exists():
            messages.error(request, f'Plate Number "{plate_no}" already exists!')
            return JsonResponse({
                'success': False,
                'message': f'Plate Number "{plate_no}" already exists!'
            }, status=400)

        # Check for duplicate Authorized Number
        if Franchise.objects.filter(authorized_no=authorized_no).exists():
            messages.error(request, f'Authorized Number "{authorized_no}" already exists!')
            return JsonResponse({
                'success': False,
                'message': f'Authorized Number "{authorized_no}" already exists!'
            }, status=400)


        # Create new Franchise record
      # Resolve tricycle FK from body_number
        tricycle_body_number = request.POST.get('tricycle_body_number', '').strip()
        tricycle = None
        if tricycle_body_number:
            try:
                tricycle = Tricycle.objects.get(body_number=tricycle_body_number)
            except Tricycle.DoesNotExist:
                pass

        franchise = Franchise.objects.create(
            tricycle=tricycle,          # ✅ this was missing
            name=name,
            denomination=denomination,
            plate_no=plate_no,
            valid_until=valid_until,
            motor_no=motor_no,
            authorized_no=authorized_no,
            chassis_no=chassis_no,
            authorized_route=authorized_route,
            purpose=purpose,
            official_receipt_no=official_receipt_no,
            date=date,
            amount_paid=amount_paid,
            municipal_treasurer=municipal_treasurer,
            status=status,
        )
        # Add success message
        messages.success(request, f'Franchise record for {name} added successfully!')

        return JsonResponse({
            'success': True,
            'message': 'Franchise record added successfully!',
            'id': franchise.id
        })

    except Exception as e:
        # Add error message
        messages.error(request, f'Error adding franchise record: {str(e)}')
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)



@require_http_methods(["POST"])
def update_franchise(request):
    try:
        # Get the franchise ID
        franchise_id = request.POST.get('id')
        
        # Get the franchise object
        franchise = Franchise.objects.get(id=franchise_id)
        
        # Get data from POST request
        plate_no = request.POST.get('plate_no')
        authorized_no = request.POST.get('authorized_no')
        
        # Check for duplicate Plate Number (excluding current record)
        if Franchise.objects.filter(plate_no=plate_no).exclude(id=franchise_id).exists():
            messages.error(request, f'Plate Number "{plate_no}" already exists!')
            return JsonResponse({
                'success': False,
                'message': f'Plate Number "{plate_no}" already exists!'
            }, status=400)

        # Check for duplicate Authorized Number (excluding current record)
        if Franchise.objects.filter(authorized_no=authorized_no).exclude(id=franchise_id).exists():
            messages.error(request, f'Authorized Number "{authorized_no}" already exists!')
            return JsonResponse({
                'success': False,
                'message': f'Authorized Number "{authorized_no}" already exists!'
            }, status=400)
        import datetime
        from decimal import Decimal
        franchise.name = request.POST.get('name')
        franchise.denomination = request.POST.get('denomination')
        franchise.plate_no = plate_no
        franchise.valid_until = datetime.date.fromisoformat(request.POST.get('valid_until'))
        franchise.motor_no = request.POST.get('motor_no')
        franchise.authorized_no = authorized_no
        franchise.chassis_no = request.POST.get('chassis_no')
        franchise.authorized_route = request.POST.get('authorized_route')
        franchise.purpose = request.POST.get('purpose')
        franchise.official_receipt_no = request.POST.get('official_receipt_no')
        franchise.date = datetime.date.fromisoformat(request.POST.get('date'))
        franchise.amount_paid = Decimal(request.POST.get('amount_paid'))
        franchise.municipal_treasurer = request.POST.get('municipal_treasurer')
        franchise.status = request.POST.get('status', 'New')

        # Save linked tricycle by body_number
        tricycle_body_number = request.POST.get('tricycle_body_number', '').strip()
        if tricycle_body_number:
            try:
                franchise.tricycle = Tricycle.objects.get(body_number=tricycle_body_number)
            except Tricycle.DoesNotExist:
                franchise.tricycle = None
        else:
            franchise.tricycle = None

        # Attach current user for the signal
        franchise._current_user = {
            'type': request.session.get('user_type'),
            'id': request.session.get('admin_id') or request.session.get('superadmin_id'),
            'name': request.session.get('full_name'),
        }

        franchise.save()

        # Add success message
        messages.success(request, f'Franchise record for {franchise.name} updated successfully!')

        return JsonResponse({
            'success': True,
            'message': 'Franchise record updated successfully!',
            'id': franchise.id
        })

    except Franchise.DoesNotExist:
        messages.error(request, 'Franchise record not found!')
        return JsonResponse({
            'success': False,
            'message': 'Franchise record not found!'
        }, status=404)

    except Exception as e:
        # Add error message
        messages.error(request, f'Error updating franchise record: {str(e)}')
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

@require_POST
def delete_franchise(request):
    try:
        franchise_id = request.POST.get('franchise_id')

        if not franchise_id:
            messages.error(request, 'Franchise ID is required')
            return JsonResponse({'success': False, 'error': 'Franchise ID is required'})

        franchise = Franchise.objects.get(id=franchise_id)

        franchise_name = franchise.name
        plate_no = franchise.plate_no

        ActivityLog.objects.create(
            action='delete',
            model_type='franchise',
            object_id=plate_no,
            object_name=franchise_name,
            description=f'Franchise record deleted: {franchise_name} (Plate No: {plate_no})',
            user_type=request.session.get('user_type'),
            user_id=request.session.get('admin_id') or request.session.get('superadmin_id'),
            user_name=request.session.get('full_name'),
        )

        franchise.delete()

        messages.success(request, f'Franchise record for "{franchise_name}" (Plate No: {plate_no}) has been deleted successfully.')
        return JsonResponse({'success': True, 'message': f'Franchise record for "{franchise_name}" has been deleted successfully.'})

    except Franchise.DoesNotExist:
        messages.error(request, 'Franchise record not found.')
        return JsonResponse({'success': False, 'error': 'Franchise record not found'})
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)})

def get_mtop(request, id):
    mtop = get_object_or_404(Mtop, id=id)

    return JsonResponse({
        "id": mtop.id,
        "name": mtop.name,
        "case_no": mtop.case_no,
        "address": mtop.address,
        "no_of_units": mtop.no_of_units,
        "route_operation": mtop.route_operation,
        "make": mtop.make,
        "motor_no": mtop.motor_no,
        "chasses_no": mtop.chasses_no,
        "plate_no": mtop.plate_no,
        "date": mtop.date.strftime("%Y-%m-%d"),
        "municipal_treasurer": mtop.municipal_treasurer,
        "officer_in_charge": mtop.officer_in_charge,
        "mayor": mtop.mayor,
    })


@require_POST
def update_mtop(request):
    import datetime
    mtop = get_object_or_404(Mtop, id=request.POST.get("id"))

    # UNIQUE FIELD CHECK (exclude self)
    if Mtop.objects.filter(case_no=request.POST["case_no"]).exclude(id=mtop.id).exists():
        messages.error(request, "Case No already exists.")
        return JsonResponse({"success": False})

    if Mtop.objects.filter(motor_no=request.POST["motor_no"]).exclude(id=mtop.id).exists():
        messages.error(request, "Motor No already exists.")
        return JsonResponse({"success": False})

    if Mtop.objects.filter(chasses_no=request.POST["chasses_no"]).exclude(id=mtop.id).exists():
        messages.error(request, "Chassis No already exists.")
        return JsonResponse({"success": False})

    # UPDATE
    mtop.name = request.POST.get("name")
    mtop.case_no = request.POST.get("case_no")
    mtop.address = request.POST.get("address")
    mtop.no_of_units = int(request.POST.get("no_of_units"))
    mtop.route_operation = request.POST.get("route_operation")
    mtop.make = request.POST.get("make")
    mtop.motor_no = request.POST.get("motor_no")
    mtop.chasses_no = request.POST.get("chasses_no")
    mtop.plate_no = request.POST.get("plate_no")
    mtop.date = datetime.date.fromisoformat(request.POST.get("date"))
    mtop.municipal_treasurer = request.POST.get("municipal_treasurer")
    mtop.officer_in_charge = request.POST.get("officer_in_charge")
    mtop.mayor = request.POST.get("mayor")

    mtop.mayor = request.POST.get("mayor")

    # Attach current user for the signal
    mtop._current_user = {
        'type': request.session.get('user_type'),
        'id': request.session.get('admin_id') or request.session.get('superadmin_id'),
        'name': request.session.get('full_name'),
    }
    mtop.save()

    messages.success(request, f"{mtop.name} successfully updated.")
    return JsonResponse({"success": True})


def admin_login(request):
    """Handle admin and superadmin login. Redirects to dashboard on success."""

    if request.session.get('admin_id') or request.session.get('superadmin_id'):
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        # First, try SuperAdmin login
        try:
            superadmin = SuperAdmin.objects.get(username=username)
            if check_password(password, superadmin.password) or superadmin.password == password:
                # Store superadmin info in session
                request.session['superadmin_id'] = superadmin.id
                request.session['user_type'] = 'superadmin'
                request.session['username'] = superadmin.username
                request.session['full_name'] = superadmin.full_name
                # Superadmins have access to all sections
                request.session['can_access_potpot_registration'] = True
                request.session['can_access_motorcycle_registration'] = True
                messages.success(request, f'Welcome back, {superadmin.full_name}!')
                return redirect('dashboard')
        except SuperAdmin.DoesNotExist:
            pass

        # If not SuperAdmin, try Admin login
        try:
            admin = Admin.objects.get(username=username)
            if check_password(password, admin.password) or admin.password == password:
                # Store admin info in session
                request.session['admin_id'] = admin.id
                request.session['user_type'] = 'admin'
                request.session['username'] = admin.username
                request.session['full_name'] = admin.full_name
                # Load admin permissions (if any) into session
                request.session['can_access_potpot_registration'] = admin.can_access_potpot_registration
                request.session['can_access_motorcycle_registration'] = admin.can_access_motorcycle_registration

                messages.success(request, f'Welcome back, {admin.full_name}!')
                return redirect('dashboard')
        except Admin.DoesNotExist:
            pass

        # If we reach here, login failed
        messages.error(request, 'Invalid username or password.')

    content = {'exclude_layout': True}
    return render(request, 'myapp/login.html', content)

def dashboard(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    
    permit_count = MayorsPermit.objects.count()
    tricycle_count = MayorsPermitTricycle.objects.count() 
     
    # Mayor's Permit (Pedicab) - Status counts
    pedicab_active = MayorsPermit.objects.filter(status='active').count()
    pedicab_inactive = MayorsPermit.objects.filter(status='inactive').count()
    pedicab_expired = MayorsPermit.objects.filter(status='expired').count()

    # Mayor's Permit (Tricycle/Motorcycle) - Status counts
    tricycle_active = MayorsPermitTricycle.objects.filter(status='active').count()
    tricycle_inactive = MayorsPermitTricycle.objects.filter(status='inactive').count()
    tricycle_expired = MayorsPermitTricycle.objects.filter(status='expired').count()
    franchise_renewed = Tricycle.objects.filter(status='Renewed').count()
    franchise_registered = Tricycle.objects.exclude(status='Inactive').count()
    franchise_expired = Tricycle.objects.filter(status='Expired').count()

   # ============ RECENT ACTIVITIES - Get from ActivityLog ============
    # Filter activities based on user permissions
    user_type = request.session.get('user_type')
    can_potpot = request.session.get('can_access_potpot_registration', False)
    can_motor = request.session.get('can_access_motorcycle_registration', False)

    # Build query based on permissions
    if user_type == 'superadmin':
        # Superadmin sees all activities
        recent_activities_raw = ActivityLog.objects.order_by('-timestamp')[:20]
    else:
        # Admin sees only activities for their permitted models
        allowed_models = []
        if can_potpot:
            allowed_models.extend(['potpot', 'idcard'])
        if can_motor:
            allowed_models.extend(['motorcycle', 'mtop', 'franchise'])
        
        if allowed_models:
            recent_activities_raw = (
                ActivityLog.objects
                .filter(model_type__in=allowed_models)
                .order_by('-timestamp')[:20]
            )
        else:
            recent_activities_raw = ActivityLog.objects.none()
    

    
    recent_activities = []
    for log in recent_activities_raw:
        # Determine icon and color based on action and model
        if log.action == 'create':
            icon = 'fa-plus-circle'
            color = 'success'
        elif log.action == 'update':
            icon = 'fa-edit'
            color = 'info'
        elif log.action == 'delete':
            icon = 'fa-trash'
            color = 'danger'
        elif log.action == 'status_change':
            icon = 'fa-sync'
            color = 'warning'
        else:
            icon = 'fa-info-circle'
            color = 'secondary'
        
        # Format the title based on model type
        model_display = dict(ActivityLog.MODEL_CHOICES).get(log.model_type, log.model_type)
        action_display = dict(ActivityLog.ACTION_CHOICES).get(log.action, log.action)
        
        activity = {
            'type': log.action,
            'model_type': log.model_type,
            'title': f'{model_display} {action_display}',
            'description': log.description,
            'timestamp': log.timestamp,
            'icon': icon,
            'color': color,
            'object_id': log.object_id,
            'updated_by': log.user_name or 'System'
        }
        recent_activities.append(activity)
    
    # Add expiring soon warnings
    expiring_soon_date = timezone.now().date() + timedelta(days=30)
    
    expiring_pedicab = MayorsPermit.objects.filter(
        expiry_date__lte=expiring_soon_date,
        expiry_date__gte=timezone.now().date(),
        status='active'
    ).count()
    
    expiring_tricycle = MayorsPermitTricycle.objects.filter(
        expiry_date__lte=expiring_soon_date,
        expiry_date__gte=timezone.now().date(),
        status='active'
    ).count()
    
    if expiring_pedicab > 0:
        recent_activities.insert(0, {
            'type': 'expiring_soon',
            'model_type': 'potpot',
            'title': 'Potpot Permits Expiring Soon',
            'description': f'{expiring_pedicab} permits expiring in the next 30 days',
            'timestamp': timezone.now(),
            'icon': 'fa-exclamation-triangle',
            'color': 'warning'
        })
    
    if expiring_tricycle > 0:
        recent_activities.insert(0, {
            'type': 'expiring_soon',
            'model_type': 'motorcycle',
            'title': 'Motorcycle Permits Expiring Soon',
            'description': f'{expiring_tricycle} permits expiring in the next 30 days',
            'timestamp': timezone.now(),
            'icon': 'fa-exclamation-triangle',
            'color': 'warning'
        })

    context = {
        'permit_count': permit_count,
        'tricycle_count': tricycle_count,
        'pedicab_active': pedicab_active,
        'pedicab_inactive': pedicab_inactive,
        'pedicab_expired': pedicab_expired,
        'tricycle_active': tricycle_active,
        'tricycle_inactive': tricycle_inactive,
        'tricycle_expired': tricycle_expired,
        'recent_activities': recent_activities,
        'franchise_renewed': franchise_renewed,
        'franchise_registered': franchise_registered,
        'franchise_expired': franchise_expired,
    }
    
    return render(request, 'myapp/dashboard.html', context)


@superadmin_required
def admin_management(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    # REMOVED .prefetch_related('permissions')
    admins = Admin.objects.all().select_related('created_by').order_by('-created_at')
    
    context = {
        'admins': admins,
    }
    return render(request, 'myapp/admin-management.html', context)


def admin_management_datatable(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    queryset = Admin.objects.select_related('created_by').order_by('-created_at')
    if search_value:
        q = (
            Q(username__icontains=search_value) |
            Q(full_name__icontains=search_value) |
            Q(email__icontains=search_value) |
            Q(created_by__username__icontains=search_value) |
            Q(permissions__can_access_potpot_registration__icontains=search_value) |
            Q(permissions__can_access_motorcycle_registration__icontains=search_value) |
            Q(created_at__icontains=search_value)
        )
        if 'inactive' in search_value.lower():
            q |= Q(is_active=False)
        elif 'active' in search_value.lower():
            q |= Q(is_active=True)
        queryset = queryset.filter(q)
    for i in range(7):
        column_search = request.GET.get(f'columns[{i}][search][value]', '')
        if column_search:
            if i == 0:   queryset = queryset.filter(username__icontains=column_search)
            elif i == 1: queryset = queryset.filter(full_name__icontains=column_search)
            elif i == 2: queryset = queryset.filter(email__icontains=column_search)
            elif i == 3: queryset = queryset.filter(created_by__username__icontains=column_search)
            elif i == 4:
                if 'potpot' in column_search.lower():
                    queryset = queryset.filter(permissions__can_access_potpot_registration=True)
                elif 'motorcycle' in column_search.lower():
                    queryset = queryset.filter(permissions__can_access_motorcycle_registration=True)
            elif i == 5:
                if column_search.lower() in ['active', 'act']:
                    queryset = queryset.filter(is_active=True)
                elif column_search.lower() in ['inactive', 'inact']:
                    queryset = queryset.filter(is_active=False)

    total_records = Admin.objects.count()
    filtered_records = queryset.count()
    admins = queryset[start:start + length]

    data = []
    for admin in admins:
        perm_potpot = admin.can_access_potpot_registration
        perm_moto = admin.can_access_motorcycle_registration

        role_display = {
            'potpot_admin': 'Potpot Admin',
            'tricycle_admin': 'Tricycle Admin',
            'both': 'Both',
        }.get(admin.role, '—')

        status_html = '<span class="badge badge-success">Active</span>' if admin.is_active else '<span class="badge badge-danger">Inactive</span>'

        created_at_str = admin.created_at.strftime('%b %d, %Y') if admin.created_at else ''

        action_html = f'''
        <div class="btn-group" role="group">
            <a href="#" class="btn btn-sm btn-info btn-admin-view"
               data-perm-potpot="{'1' if perm_potpot else '0'}"
               data-perm-moto="{'1' if perm_moto else '0'}">
                <i class="fas fa-eye"></i>
            </a>
            <a href="#" class="btn btn-sm btn-warning btn-admin-edit"
               data-admin-id="{admin.id}"
               data-username="{admin.username}"
               data-full-name="{admin.full_name}"
               data-email="{admin.email}"
               data-created-by="{admin.created_by.username if admin.created_by else ''}"
               data-created-at="{created_at_str}"
               data-status="{'1' if admin.is_active else '0'}"
               data-role="{admin.role}"
               data-perm-potpot="{'1' if perm_potpot else '0'}"
               data-perm-moto="{'1' if perm_moto else '0'}">
                <i class="fas fa-edit"></i>
            </a>
            <a href="#" class="btn btn-sm btn-danger btn-admin-delete"
               data-admin-id="{admin.id}">
                <i class="fas fa-trash"></i>
            </a>
        </div>
        '''

        data.append([
            admin.username,                                          # 0 - Username
            admin.full_name,                                         # 1 - Full Name
            admin.email,                                             # 2 - Email
            admin.created_by.username if admin.created_by else '—', # 3 - Created By
            role_display,# 4 - Permissions
            status_html,                                             # 5 - Status
            action_html,                                             # 6 - Action
            str(admin.id),                                           # 7 - Hidden ID
            created_at_str,                                          # 8 - Hidden created_at
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })

@require_POST
@superadmin_required
def add_admin(request):
    super_admin_id = request.session.get('superadmin_id')

    if not super_admin_id:
        messages.error(request, "Unauthorized access.")
        return redirect('admin-management')

    super_admin = SuperAdmin.objects.get(id=super_admin_id)

    username = request.POST.get('username')
    full_name = request.POST.get('full_name')
    email = request.POST.get('email')
    password = request.POST.get('password')

    # Permissions
    role = request.POST.get('role', 'potpot_admin')

    if Admin.objects.filter(username=username).exists():
        messages.error(request, "Username already exists.")
        return redirect('admin-management')

    if Admin.objects.filter(email=email).exists():
        messages.error(request, "Email already exists.")
        return redirect('admin-management')

    try:
        # ← Only one create, no more AdminPermission.objects.create()
        Admin.objects.create(
            username=username,
            full_name=full_name,
            email=email,
            password=make_password(password),
            role=role,              # ← NEW
            created_by=super_admin
        )

        messages.success(request, "Admin created successfully.")

    except Exception as e:
        messages.error(request, f"Error creating admin: {str(e)}")

    return redirect('admin-management')


# @superadmin_required
# def get_admin_details(request, admin_id):
#     """
#     Returns JSON with all admin details including permissions and created_by info
#     """
#     admin = get_object_or_404(Admin.objects.select_related('created_by', 'permissions'), id=admin_id)

#     data = {
#         "id": admin.id,
#         "username": admin.username,
#         "full_name": admin.full_name,
#         "email": admin.email,
#         "is_active": admin.is_active,
#         "created_by": admin.created_by.full_name if admin.created_by else None,
#         "created_at": admin.created_at.strftime("%b %d, %Y"),
#         "permissions": {
#             "can_access_potpot_registration": admin.permissions.can_access_potpot_registration,
#             "can_access_motorcycle_registration": admin.permissions.can_access_motorcycle_registration,
#         }
#     }

#     return JsonResponse(data)

@require_POST
@superadmin_required
def update_admin(request):
    """Handle updates to an Admin and their permissions."""
    super_admin_id = request.session.get('superadmin_id')
    if not super_admin_id:
        messages.error(request, "Unauthorized access.")
        return redirect('admin-management')

    super_admin = SuperAdmin.objects.get(id=super_admin_id)

    admin_id = request.POST.get('admin_id')
    if not admin_id:
        messages.error(request, "Missing admin identifier.")
        return redirect('admin-management')

    admin = get_object_or_404(Admin, id=admin_id)

    username = request.POST.get('username', '').strip()
    full_name = request.POST.get('full_name', '').strip()
    email = request.POST.get('email', '').strip()
    status = request.POST.get('status')
    is_active = True if status == '1' else False

    # Validate uniqueness if changed
    if username and Admin.objects.filter(username=username).exclude(id=admin.id).exists():
        messages.error(request, "Username already exists.")
        return redirect('admin-management')

    if email and Admin.objects.filter(email=email).exclude(id=admin.id).exists():
        messages.error(request, "Email already exists.")
        return redirect('admin-management')

    admin.username = username or admin.username
    admin.full_name = full_name or admin.full_name
    admin.email = email or admin.email
    admin.is_active = is_active

    password = request.POST.get('password', '')
    if password:
        admin.password = make_password(password)

    admin.role = request.POST.get('role', admin.role)
    admin.save()

    messages.success(request, "Admin updated successfully.")
    return redirect('admin-management')


def admin_list(request):
   
    return render(request, 'myapp/admin-list.html')

def admin_logout(request):
    """Clear admin and superadmin session and redirect to login."""
    
    # Clear all session keys used in login
    session_keys = ['admin_id', 'superadmin_id', 'user_type', 'username', 'full_name', 'can_access_potpot_registration', 'can_access_motorcycle_registration']
    
    for key in session_keys:
        if key in request.session:
            del request.session[key]
    
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


def mayors_permit(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    permits = MayorsPermit.objects.all()
    return render(request, 'myapp/mayors-permit.html', {'permits': permits})



def mayors_permit_datatable(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    """Server-side processing endpoint for DataTables"""
    
    # DataTables parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    
    # Base queryset
    queryset = MayorsPermit.objects.all()
    
    # Global search
    if search_value:
        queryset = queryset.filter(
            Q(control_no__icontains=search_value) |
            Q(name__icontains=search_value) |
            Q(address__icontains=search_value) |
            Q(business_name__icontains=search_value) |
            Q(or_no__icontains=search_value) |
            Q(motorized_operation__icontains=search_value) |
            Q(issue_date__icontains=search_value) |
            Q(expiry_date__icontains=search_value) |
            Q(issued_at__icontains=search_value) |
            Q(mayor__icontains=search_value) |
            Q(quarter__icontains=search_value) |
            Q(status__icontains=search_value) |
            Q(amount_paid__icontains=search_value)
        )
    
    # Column-specific search
    for i in range(14):  # 14 columns total
        column_search = request.GET.get(f'columns[{i}][search][value]', '')
        if column_search:
            if i == 0:  # Control No
                queryset = queryset.filter(control_no__icontains=column_search)
            elif i == 1:  # Status - exact match
                queryset = queryset.filter(status__iexact=column_search)
            elif i == 2:  # Name
                queryset = queryset.filter(name__icontains=column_search)
            elif i == 3:  # Address
                queryset = queryset.filter(address__icontains=column_search)
            elif i == 4:  # Business Name
                queryset = queryset.filter(business_name__icontains=column_search)
            elif i == 5:  # Motorized Operation
                queryset = queryset.filter(motorized_operation__icontains=column_search)
            elif i == 6:  # OR No
                queryset = queryset.filter(or_no__icontains=column_search)
            elif i == 7:  # Amount Paid
                queryset = queryset.filter(amount_paid__icontains=column_search)
            elif i == 8:  # Issue Date
                queryset = queryset.filter(issue_date__icontains=column_search)
            elif i == 9:  # Expiry Date
                queryset = queryset.filter(expiry_date__icontains=column_search)
            elif i == 10:  # Issued At
                queryset = queryset.filter(issued_at__icontains=column_search)
            elif i == 11:  # Mayor
                queryset = queryset.filter(mayor__icontains=column_search)
            elif i == 12:  # Quarter
                queryset = queryset.filter(quarter__icontains=column_search)
    
    # Total records
    total_records = MayorsPermit.objects.count()
    filtered_records = queryset.count()
    
    # Pagination
    permits = queryset[start:start + length]
    
    # Format data
    data = []
    for permit in permits:
        # Status badge HTML
      
        
        # Generate URLs using reverse()
        print_url = reverse('print-mayors-permit', args=[permit.id])
        history_url = reverse('mayors-permit-history', args=[permit.id])
        
      # Action buttons HTML - MODIFIED HISTORY BUTTON
        action_html = f'''
        <div class="btn-group" role="group">
            <a href="#" class="btn btn-sm btn-primary btn-view" data-id="{permit.id}" title="View">
                <i class="fas fa-eye"></i>
            </a>
            <a href="{print_url}" target="_blank" class="btn btn-sm btn-secondary" title="Print">
                <i class="fas fa-print"></i>
            </a>
            <a href="#" class="btn btn-sm btn-info btn-update" data-id="{permit.id}" title="Update">
                <i class="fas fa-edit"></i>
            </a>
            <a href="#" class="btn btn-sm btn-warning btn-history" data-id="{permit.id}" title="Status history">
                <i class="fas fa-history"></i>
            </a>
        </div>
        '''
        
       
        data.append([
            permit.control_no,
            permit.status,  # ← Send raw status text instead of HTML
            permit.name,
            permit.address,
            permit.business_name or '',
            permit.motorized_operation or '',
            permit.or_no,
            '{:,.2f}'.format(permit.amount_paid),
            (permit.issue_date.strftime('%b-') + str(permit.issue_date.day) + permit.issue_date.strftime('-%Y')),
            (permit.expiry_date.strftime('%b-') + str(permit.expiry_date.day) + permit.expiry_date.strftime('-%Y')),
            permit.issued_at or '',
            permit.mayor or '',
            permit.get_quarter_display(),
            action_html,
            permit.id,
            permit.status  # Hidden column for raw status value
        ])
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })

def get_permit_history(request, permit_id):
    """Get history data for a specific permit"""
    try:
        permit = MayorsPermit.objects.get(id=permit_id)
        
        # Fetch actual history records from database
        history_records = MayorsPermitHistory.objects.filter(permit=permit).order_by('-activated_at')
        
        history_list = []
        for record in history_records:
            history_list.append({
                'date':localtime(record.activated_at).strftime('%Y-%m-%d %I:%M %p'
),
                'old_status': record.previous_status.capitalize(),
                'new_status': record.new_status.capitalize(),
                'changed_by': record.updated_by_name or 'System',
                'reason': record.remarks or ''
            })
        
        history_data = {
            'permit_info': {
                'control_no': permit.control_no,
                'name': permit.name,
                'current_status': permit.status,
            },
            'history': history_list
        }
        
        return JsonResponse({'success': True, 'data': history_data})
        
    except MayorsPermit.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Permit not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
def add_mayors_permit(request):
    if request.method == "POST":
        control_no = request.POST.get("control_no")
        name = request.POST.get("name")
        address = request.POST.get("address")
        motorized_operation = request.POST.get("motorized_operation")
        business_name = request.POST.get("business_name")
        expiry_date = request.POST.get("expiry_date")
        amount_paid = request.POST.get("amount_paid")
        or_no = request.POST.get("or_no")
        issue_date = request.POST.get("issue_date")
        issued_at = request.POST.get("issued_at")
        mayor = request.POST.get("mayor")
        quarter = request.POST.get("quarter")
        status = request.POST.get("status")

        if MayorsPermit.objects.filter(control_no=control_no).exists():
            return JsonResponse({"success": False, "error": f"Control No '{control_no}' is already used."})

        if MayorsPermit.objects.filter(or_no=or_no).exists():
            return JsonResponse({"success": False, "error": f"OR No '{or_no}' is already used."})

        MayorsPermit.objects.create(
            control_no=control_no,
            name=name,
            address=address,
            motorized_operation=motorized_operation,
            business_name=business_name,
            expiry_date=expiry_date,
            amount_paid=amount_paid,
            or_no=or_no,
            issue_date=issue_date,
            issued_at=issued_at,
            mayor=mayor,
            quarter=quarter,
            status=status
        )
        messages.success(request, "Mayor's Permit created successfully!")
        return JsonResponse({"success": True, "message": "Mayor's Permit created successfully!"})

    return JsonResponse({"success": False, "error": "Invalid request method."})


def update_mayors_permit(request, permit_id):

    permit = get_object_or_404(MayorsPermit, id=permit_id)
    old_status = permit.status
    # Extract POST datar
    control_no = request.POST.get("control_no")
    name = request.POST.get("name")
    address = request.POST.get("address")
    business_name = request.POST.get("business_name")
    motorized_operation = request.POST.get("motorized_operation")
    or_no = request.POST.get("or_no")
    amount_paid = request.POST.get("amount_paid")
    issue_date = parse_date(request.POST.get("issue_date"))
    expiry_date = parse_date(request.POST.get("expiry_date"))
    issued_at = request.POST.get("issued_at")
    mayor = request.POST.get("mayor")
    quarter = request.POST.get("quarter")
    status = request.POST.get("status")

    # Update fields
    permit.control_no = control_no
    permit.name = name
    permit.address = address
    permit.business_name = business_name
    permit.motorized_operation = motorized_operation
    permit.or_no = or_no
    permit.amount_paid = amount_paid
    permit.issue_date = issue_date
    permit.expiry_date = expiry_date
    permit.issued_at = issued_at
    permit.mayor = mayor
    permit.quarter = quarter
    permit.status = status

    permit.save()
  # ✅ RECORD STATUS CHANGES WITH USER INFO
    if old_status != permit.status:
        # Get user info from session
        user_type = request.session.get('user_type')
        user_id = request.session.get('superadmin_id') if user_type == 'superadmin' else request.session.get('admin_id')
        user_name = request.session.get('full_name')
        
        MayorsPermitHistory.objects.create(
            permit=permit,
            previous_status=old_status,
            new_status=permit.status,
            remarks=f"Status changed from {old_status} to {permit.status}",
            updated_by_type=user_type,
            updated_by_id=user_id,
            updated_by_name=user_name
        )
    messages.success(request, f"Permit for **{permit.name}** updated successfully!")

    # Return JSON for AJAX success
    return JsonResponse({"success": True, "message": "Permit updated successfully!"})

def id_cards(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    cards = IDCard.objects.all()
    return render(request, 'myapp/id-cards.html', {'cards': cards})


def id_cards_datatable(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    queryset = IDCard.objects.all()

    # Global search
    if search_value:
        # Map display text to stored gender code
        gender_map = {'male': 'M', 'female': 'F', 'other': 'O'}
        gender_code = gender_map.get(search_value.lower(), search_value)

        queryset = queryset.filter(
            Q(name__icontains=search_value) |
            Q(id_number__icontains=search_value) |
            Q(address__icontains=search_value) |
            Q(or_number__icontains=search_value) |
            Q(gender__icontains=gender_code) |        # ← use gender_code here
            Q(date_of_birth__icontains=search_value) |
            Q(height__icontains=search_value) |
            Q(weight__icontains=search_value) |
            Q(date_issued__icontains=search_value) |
            Q(expiration_date__icontains=search_value)
        )

    # Column-specific search
    column_map = {
        0: None,          # Image - not searchable
        1: 'id_number',
        2: 'name',
        3: 'gender',
        4: 'date_of_birth',
        5: 'address',
        6: 'or_number',
        7: 'height',
        8: 'weight',
        9: 'date_issued',
        10: 'expiration_date',
    }

    for i, field in column_map.items():
        if field is None:
            continue
        col_search = request.GET.get(f'columns[{i}][search][value]', '')
        if col_search:
            # Remap gender display value to stored code
            if field == 'gender':
                gender_map = {'male': 'M', 'female': 'F', 'other': 'O'}
                col_search = gender_map.get(col_search.lower(), col_search)
            queryset = queryset.filter(**{f'{field}__icontains': col_search})

    total_records = IDCard.objects.count()
    filtered_records = queryset.count()

    cards = queryset[start:start + length]

    data = []
    for card in cards:
        image_html = ''
        if card.image:
            image_html = f'<img src="{card.image.url}" class="id-thumb" style="cursor:pointer;" onclick="showImagePreview(\'{card.image.url}\')">'
        else:
            image_html = '<span class="text-muted">-</span>'

        image_url = card.image.url if card.image else ''
        gender_display = card.get_gender_display()
        dob = (card.date_of_birth.strftime('%b-') + str(card.date_of_birth.day) + card.date_of_birth.strftime('-%Y')) if card.date_of_birth else ''
        date_issued = (card.date_issued.strftime('%b-') + str(card.date_issued.day) + card.date_issued.strftime('-%Y')) if card.date_issued else ''
        expiry = (card.expiration_date.strftime('%b-') + str(card.expiration_date.day) + card.expiration_date.strftime('-%Y')) if card.expiration_date else ''
        height = str(card.height) if card.height else ''
        weight = str(card.weight) if card.weight else ''
        or_number = card.or_number or ''

        action_html = f'''
        <div class="btn-group btn-group-sm" role="group">
            <button type="button" class="btn btn-info btn-view-card"
                data-image="{image_url}"
                data-name="{card.name}"
                data-idnumber="{card.id_number}"
                data-gender="{gender_display}"
                data-dob="{dob}"
                data-address="{card.address}"
                data-ornumber="{or_number}"
                data-height="{height}"
                data-weight="{weight}"
                data-dateissued="{date_issued}"
                data-expiry="{expiry}"
                title="View">
                <i class="fas fa-eye"></i>
            </button>
            <button type="button" class="btn btn-warning btn-update-card"
                data-id="{card.id}"
                data-image="{image_url}"
                data-name="{card.name}"
                data-idnumber="{card.id_number}"
                data-gender="{card.gender}"
                data-dob="{dob}"
                data-address="{card.address}"
                data-ornumber="{or_number}"
                data-height="{height}"
                data-weight="{weight}"
                data-dateissued="{date_issued}"
                data-expiry="{expiry}"
                title="Update">
                <i class="fas fa-edit"></i>
            </button>
            <button type="button" class="btn btn-danger btn-delete-card" data-id="{card.id}" title="Delete">
                <i class="fas fa-trash"></i>
            </button>
            <button type="button" class="btn btn-secondary btn-print-card"
                data-image="{image_url}"
                data-name="{card.name}"
                data-idnumber="{card.id_number}"
                data-address="{card.address}"
                data-dob="{dob}"
                data-gender="{gender_display}"
                data-ornumber="{or_number}"
                data-height="{height}"
                data-weight="{weight}"
                data-dateissued="{date_issued}"
                data-expiry="{expiry}"
                data-dmonate="DM Onate"
                title="Print">
                <i class="fas fa-print"></i>
            </button>
        </div>
        '''

        data.append([
            image_html,       # 0 - Image
            card.id_number,   # 1
            card.name,        # 2
            gender_display,   # 3
            dob,              # 4
            card.address,     # 5
            or_number,        # 6
            height,           # 7
            weight,           # 8
            date_issued,      # 9
            expiry,           # 10
            action_html,      # 11
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data,
    })

# ============ ID CARD IMPORT/EXPORT WITH IMAGES ============

def export_idcards_with_images(request):
    """Export all ID Cards with their images as a ZIP file"""
    try:
        # Create a BytesIO object to store the zip file
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Create CSV with ID card data
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer)
            
            # Write header
            writer.writerow([
                'id', 'name', 'id_number', 'address', 'date_of_birth', 'gender',
                'or_number', 'height', 'weight', 'date_issued', 'expiration_date', 'image_filename'
            ])
            
            # Get all ID cards
            cards = IDCard.objects.all()
            
            for card in cards:
                image_filename = ''
                
                # Add image to zip if it exists
                if card.image:
                    # Get image path and filename
                    image_path = card.image.path
                    image_name = os.path.basename(image_path)
                    image_filename = f"images/{image_name}"
                    
                    # Add image to zip
                    try:
                        zip_file.write(image_path, image_filename)
                    except Exception as e:
                        print(f"Error adding image {image_name}: {e}")
                
                # Write card data to CSV
                writer.writerow([
                    card.id,
                    card.name,
                    card.id_number,
                    card.address,
                    card.date_of_birth.strftime("%Y-%m-%d") if card.date_of_birth else '',
                    card.gender,
                    card.or_number or '',
                    card.height or '',
                    card.weight or '',
                    card.date_issued.strftime("%Y-%m-%d") if card.date_issued else '',
                    card.expiration_date.strftime("%Y-%m-%d") if card.expiration_date else '',
                    image_filename
                ])
            
            # Add CSV to zip
            zip_file.writestr('id_cards.csv', csv_buffer.getvalue())
        
        # Prepare response
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="id_cards_export.zip"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Error exporting ID Cards: {str(e)}")
        return redirect('id-cards')


@require_POST
def import_idcards_with_images(request):
    """Import ID Cards from ZIP file containing CSV and images"""
    try:
        uploaded_file = request.FILES.get('zip_file')
        
        if not uploaded_file:
            messages.error(request, "No file uploaded.")
            return redirect('id-cards')
        
        if not uploaded_file.name.lower().endswith('.zip'):
            messages.error(request, "Please upload a ZIP file.")
            return redirect('id-cards')
        
        success_count = 0
        errors = []
        
        # Create a temporary directory to extract files
        temp_dir = tempfile.mkdtemp()
        
        try:
            with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Read the CSV file
            csv_path = os.path.join(temp_dir, 'id_cards.csv')
            
            if not os.path.exists(csv_path):
                messages.error(request, "CSV file not found in ZIP.")
                return redirect('id-cards')
            
            with open(csv_path, 'r', encoding='utf-8') as csv_file:
                reader = csv.DictReader(csv_file)
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Parse dates
                        date_of_birth = None
                        if row.get('date_of_birth'):
                            date_of_birth = datetime.strptime(row['date_of_birth'], '%Y-%m-%d').date()
                        
                        date_issued = None
                        if row.get('date_issued'):
                            date_issued = datetime.strptime(row['date_issued'], '%Y-%m-%d').date()
                        
                        expiration_date = None
                        if row.get('expiration_date'):
                            expiration_date = datetime.strptime(row['expiration_date'], '%Y-%m-%d').date()
                        
                        # Handle image - READ INTO MEMORY BEFORE TEMP CLEANUP
                        image_name = None
                        image_data = None
                        if row.get('image_filename'):
                            image_path = os.path.join(temp_dir, row['image_filename'])
                            if os.path.exists(image_path):
                                with open(image_path, 'rb') as img:
                                    image_data = img.read()
                                    image_name = os.path.basename(image_path)
                        
                        # Get or create card
                        card, created = IDCard.objects.update_or_create(
                            id_number=row['id_number'],
                            defaults={
                                'name': row['name'],
                                'address': row['address'],
                                'date_of_birth': date_of_birth,
                                'gender': row['gender'],
                                'or_number': row.get('or_number', ''),
                                'height': float(row['height']) if row.get('height') else None,
                                'weight': float(row['weight']) if row.get('weight') else None,
                                'date_issued': date_issued,
                                'expiration_date': expiration_date,
                            }
                        )
                        
                        # Save image if it exists (AFTER temp data is in memory)
                        if image_data and image_name:
                            if card.image:
                                card.image.delete(save=False)
                            card.image.save(image_name, BytesIO(image_data), save=True)
                        
                        success_count += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
        
        finally:
            # Clean up temp directory AFTER all image data has been read into memory
            shutil.rmtree(temp_dir)
        
        if success_count:
            messages.success(request, f"{success_count} ID Cards imported successfully!")
        
        if errors:
            error_msg = "Some rows could not be imported:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                error_msg += f"\n... and {len(errors) - 10} more errors."
            messages.warning(request, error_msg)
        
        return redirect('id-cards')
        
    except Exception as e:
        messages.error(request, f"Error importing ID Cards: {str(e)}")
        return redirect('id-cards')
    
def update_idcard(request):
    import datetime
    from decimal import Decimal
    """Handle ID Card update"""
    if request.method == 'POST':
        try:
            card_id = request.POST.get('card_id')
            card = get_object_or_404(IDCard, id=card_id)
            
            # Update fields
            card.name = request.POST.get('name')
            card.id_number = request.POST.get('id_number')
            card.address = request.POST.get('address')
            card.gender = request.POST.get('gender')
            card.or_number = request.POST.get('or_number', '')

           

            if request.POST.get('date_of_birth'):
                card.date_of_birth = datetime.date.fromisoformat(request.POST.get('date_of_birth'))

            if request.POST.get('height'):
                card.height = Decimal(request.POST.get('height'))

            if request.POST.get('weight'):
                card.weight = Decimal(request.POST.get('weight'))

            if request.POST.get('date_issued'):
                card.date_issued = datetime.date.fromisoformat(request.POST.get('date_issued'))

            if request.POST.get('expiration_date'):
                card.expiration_date = datetime.date.fromisoformat(request.POST.get('expiration_date'))
            
            if request.FILES.get('image'):
                if card.image:
                    card.image.delete(save=False)
                image = request.FILES['image']
                card.image.save(image.name, image, save=False)  # save=False because card.save() is called below
            
            # Attach current user info for the signal to use
            user_type = request.session.get('user_type')  # 'admin' or 'superadmin'
            user_id = request.session.get('user_id')
            user_name = request.session.get('full_name')  # or 'full_name', adjust to your session keys

            card._current_user = {
                'type': user_type,
                'id': user_id,
                'name': user_name,
            }
            
            card.save()
            
            messages.success(request, 'ID Card updated successfully!')
            return JsonResponse({'status': 'success', 'message': 'ID Card updated successfully!'})
            
        except Exception as e:
            messages.error(request, f'Error updating ID Card: {str(e)}')
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def add_idcard(request):
    if request.method == "POST":
        name = request.POST.get("name")
        address = request.POST.get("address")
        id_number = request.POST.get("id_number")
        date_of_birth = request.POST.get("date_of_birth")
        gender = request.POST.get("gender")
        or_number = request.POST.get("or_number")
        height = request.POST.get("height")
        weight = request.POST.get("weight")
        date_issued = request.POST.get("date_issued")
        expiration_date = request.POST.get("expiration_date")
        image = request.FILES.get("image")

        # Create card first WITHOUT image
        card = IDCard.objects.create(
            name=name,
            address=address,
            id_number=id_number,
            date_of_birth=date_of_birth,
            gender=gender,
            or_number=or_number,
            height=height if height else None,
            weight=weight if weight else None,
            date_issued=date_issued,
            expiration_date=expiration_date,
        )

        # Then save image separately so Cloudinary storage backend handles it properly
        if image:
            card.image.save(image.name, image, save=True)

        messages.success(request, "ID Card added successfully!")
        return JsonResponse({"success": True})

    return JsonResponse({"success": False})


def permit_renewal(request):
    permits = MayorsPermit.objects.all()
    return render(request, 'myapp/permit-renewal.html', {'permits': permits})

def permit_detail_api(request, control_no):
    permit = get_object_or_404(MayorsPermit, control_no=control_no)

    data = {
        "control_no": permit.control_no,
        "name": permit.name,
        "address": permit.address,
        "business_name": permit.business_name,
        "status": permit.status,
        "issue_date": permit.issue_date.strftime("%Y-%m-%d"),
        "expiry_date": permit.expiry_date.strftime("%Y-%m-%d"),
        "mayor": permit.mayor,
        "quarter": permit.get_quarter_display(),
    }

    return JsonResponse(data, json_dumps_params={'indent': 4})

def print_mayors_permit(request, pk):
    permit = get_object_or_404(MayorsPermit, pk=pk)
    return render(request, "myapp/mayors-permit-print.html", {"permit": permit})


def mtop(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    mtops = Mtop.objects.all()
    return render(request, 'myapp/mtop.html', {'mtops': mtops})

def mtop_datatable(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    """Server-side processing endpoint for MTOP DataTables"""
    
    # DataTables parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    
    # Base queryset
    queryset = Mtop.objects.all()
    
    # Global search
    if search_value:
        queryset = queryset.filter(
            Q(name__icontains=search_value) |
            Q(case_no__icontains=search_value) |
            Q(address__icontains=search_value) |
            Q(route_operation__icontains=search_value) |
            Q(make__icontains=search_value) |
            Q(motor_no__icontains=search_value) |
            Q(chasses_no__icontains=search_value) |
            Q(plate_no__icontains=search_value) |
            Q(municipal_treasurer__icontains=search_value) |
            Q(officer_in_charge__icontains=search_value) |
            Q(mayor__icontains=search_value) |
            Q(date__icontains=search_value)
        )
    
    # Column-specific search (matching your 14 columns)
    for i in range(14):
        column_search = request.GET.get(f'columns[{i}][search][value]', '')
        if column_search:
            if i == 0:  # Name
                queryset = queryset.filter(name__icontains=column_search)
            elif i == 1:  # Case No
                queryset = queryset.filter(case_no__icontains=column_search)
            elif i == 2:  # Address
                queryset = queryset.filter(address__icontains=column_search)
            elif i == 3:  # No. of Units
                queryset = queryset.filter(no_of_units__icontains=column_search)
            elif i == 4:  # Route Operation
                queryset = queryset.filter(route_operation__icontains=column_search)
            elif i == 5:  # Make
                queryset = queryset.filter(make__icontains=column_search)
            elif i == 6:  # Motor No
                queryset = queryset.filter(motor_no__icontains=column_search)
            elif i == 7:  # Chassis No
                queryset = queryset.filter(chasses_no__icontains=column_search)
            elif i == 8:  # Plate No
                queryset = queryset.filter(plate_no__icontains=column_search)
            elif i == 9:  # Date
                queryset = queryset.filter(date__icontains=column_search)
            elif i == 10:  # Municipal Treasurer
                queryset = queryset.filter(municipal_treasurer__icontains=column_search)
            elif i == 11:  # Officer in Charge
                queryset = queryset.filter(officer_in_charge__icontains=column_search)
            elif i == 12:  # Mayor
                queryset = queryset.filter(mayor__icontains=column_search)
    
    # Total records
    total_records = Mtop.objects.count()
    filtered_records = queryset.count()
    
    # Pagination
    mtops = queryset[start:start + length]
    
    # Format data
    data = []
    for m in mtops:
        # Generate the print URL dynamically using reverse()
        print_url = reverse('mtop_print', args=[m.id])
        
        action_html = f'''
        <div class="btn-group" role="group" aria-label="actions">
            <a href="{print_url}" class="btn btn-sm btn-primary" title="Print" target="_blank">
                <i class="fas fa-print"></i>
            </a>
            <button class="btn btn-sm btn-warning edit-btn" 
                data-id="{m.id}" 
                data-toggle="modal" 
                data-target="#editRecordModal">
                <i class="fas fa-edit"></i>
            </button>
            <button class="btn btn-sm btn-danger btn-delete-mtop" title="Delete"
                data-id="{m.id}"
                data-case-no="{m.case_no}"
                data-name="{m.name}">
                <i class="fas fa-trash"></i>
            </button>
        </div>
        '''
        
        data.append([
            m.name,                                      # 0
            m.case_no,                                   # 1
            m.address,                                   # 2
            m.no_of_units,                              # 3
            m.route_operation,                          # 4
            m.make,                                      # 5
            m.motor_no,                                  # 6
            m.chasses_no,                               # 7
            m.plate_no,                                  # 8
            m.date.strftime('%b-') + str(m.date.day) + m.date.strftime('-%Y'),  # 9             # 9
            m.municipal_treasurer,                      # 10
            m.officer_in_charge,                        # 11
            m.mayor,                                     # 12
            action_html                                  # 13
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })

@require_POST
def delete_mtop(request):
    try:
        mtop_id = request.POST.get('mtop_id')

        if not mtop_id:
            messages.error(request, 'MTOP ID is required')
            return JsonResponse({'success': False, 'error': 'MTOP ID is required'})

        mtop = Mtop.objects.get(id=mtop_id)

        case_no = mtop.case_no
        name = mtop.name

        ActivityLog.objects.create(
            action='delete',
            model_type='mtop',
            object_id=case_no,
            object_name=name,
            description=f'MTOP record deleted: {name} (Case No: {case_no})',
            user_type=request.session.get('user_type'),
            user_id=request.session.get('admin_id') or request.session.get('superadmin_id'),
            user_name=request.session.get('full_name'),
        )

        mtop.delete()

        messages.success(request, f'MTOP record "{case_no}" for {name} has been deleted successfully.')
        return JsonResponse({'success': True, 'message': f'MTOP record "{case_no}" has been deleted successfully.'})

    except Mtop.DoesNotExist:
        messages.error(request, 'MTOP record not found.')
        return JsonResponse({'success': False, 'error': 'MTOP record not found'})
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)})

def import_mtop(request):
    """Import MTOP records from CSV or Excel file"""
    if request.method != "POST":
        messages.error(request, 'Invalid request method')
        return redirect('mtop')
    
    # Check if file exists in request
    uploaded_file = request.FILES.get('import_file')
    if not uploaded_file:
        messages.error(request, "No file uploaded.")
        return redirect('mtop')
    
    file_name = uploaded_file.name.lower()
    is_excel = file_name.endswith(('.xlsx', '.xls'))
    is_csv = file_name.endswith('.csv')
    
    if not (is_csv or is_excel):
        messages.error(request, "Please upload a CSV or Excel file (.csv, .xlsx, .xls)")
        return redirect('mtop')
    
    errors = []
    success_count = 0
    
    # Expected header columns
    expected_header = ['name', 'case_no', 'address', 'no_of_units', 'route_operation',
                      'make', 'motor_no', 'chasses_no', 'plate_no', 'date',
                      'municipal_treasurer', 'officer_in_charge', 'mayor']
    
    def normalize_header(header):
        """Convert header to lowercase, replace spaces/periods with underscores, and fix common variations"""
        normalized = []
        for h in header:
            # Convert to lowercase, strip whitespace
            h = h.strip().lower()
            # Replace spaces and periods with underscores
            h = h.replace(' ', '_').replace('.', '')
            # Fix common spelling variations
            h = h.replace('chassis', 'chasses')  # chassis -> chasses
            normalized.append(h)
        return normalized
    
    try:
        if is_excel:
            # Handle Excel file
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows:
                messages.error(request, "Excel file is empty.")
                return redirect('mtop')
            
            # Validate and normalize header
            raw_header = [str(cell).strip() if cell else '' for cell in rows[0]]
            normalized_header = normalize_header(raw_header)
            
            if normalized_header != expected_header:
                messages.error(request, f"Invalid Excel header. Expected: {', '.join(expected_header)}. Got: {', '.join(normalized_header)}")
                return redirect('mtop')
            
            # Skip header row
            data_rows = rows[1:]
            
            for i, row in enumerate(data_rows, start=2):
                if not row or all(cell is None for cell in row):
                    continue
                
                if len(row) < len(expected_header):
                    errors.append(f"Row {i}: Incorrect number of columns ({len(row)}). Expected {len(expected_header)}.")
                    continue
                
                try:
                    # Handle date parsing for Excel - keep original date type
                    def parse_excel_date(date_val):
                        if isinstance(date_val, datetime):
                            return date_val.date()
                        elif isinstance(date_val, str):
                            return datetime.strptime(date_val.strip(), "%Y-%m-%d").date()
                        else:
                            return date_val
                    
                    date_obj = parse_excel_date(row[9])
                    
                    try:
                        no_of_units = int(float(row[3])) if row[3] else 1
                    except ValueError:
                        raise ValueError(f"Invalid no_of_units '{row[3]}'")
                    
                    Mtop.objects.update_or_create(
                        case_no=str(row[1]).strip() if row[1] else '',
                        defaults={
                            "name": str(row[0]).strip() if row[0] else '',
                            "address": str(row[2]).strip() if row[2] else '',
                            "no_of_units": no_of_units,
                            "route_operation": str(row[4]).strip() if row[4] else '',
                            "make": str(row[5]).strip() if row[5] else '',
                            "motor_no": str(row[6]).strip() if row[6] else '',
                            "chasses_no": str(row[7]).strip() if row[7] else '',
                            "plate_no": str(row[8]).strip() if row[8] else '',
                            "date": date_obj,
                            "municipal_treasurer": str(row[10]).strip() if row[10] else '',
                            "officer_in_charge": str(row[11]).strip() if row[11] else '',
                            "mayor": str(row[12]).strip() if row[12] else '',
                        },
                    )
                    success_count += 1
                
                except ValueError as ve:
                    errors.append(f"Row {i}: Value error - {ve}")
                except Exception as e:
                    errors.append(f"Row {i}: {e}")
        
        else:
            # Handle CSV file
            file_data = uploaded_file.read().decode("utf-8-sig").splitlines()
            reader = csv.reader(file_data)
            raw_header = next(reader)
            
            # Validate and normalize header columns
            normalized_header = normalize_header(raw_header)
            
            if normalized_header != expected_header:
                messages.error(request, f"Invalid CSV header. Expected: {', '.join(expected_header)}. Got: {', '.join(normalized_header)}")
                return redirect('mtop')
            
            for i, row in enumerate(reader, start=2):
                # Strip whitespace from all columns
                row = [c.strip() for c in row]
                
                if len(row) != len(expected_header):
                    errors.append(f"Row {i}: Incorrect number of columns ({len(row)}). Expected {len(expected_header)}.")
                    continue
                
                try:
                    date_obj = datetime.strptime(row[9], "%Y-%m-%d").date()
                    
                    try:
                        no_of_units = int(row[3])
                    except ValueError:
                        raise ValueError(f"Invalid no_of_units '{row[3]}'")
                    
                    Mtop.objects.update_or_create(
                        case_no=row[1],
                        defaults={
                            "name": row[0],
                            "address": row[2],
                            "no_of_units": no_of_units,
                            "route_operation": row[4],
                            "make": row[5],
                            "motor_no": row[6],
                            "chasses_no": row[7],
                            "plate_no": row[8],
                            "date": date_obj,
                            "municipal_treasurer": row[10],
                            "officer_in_charge": row[11],
                            "mayor": row[12],
                        },
                    )
                    success_count += 1
                
                except ValueError as ve:
                    errors.append(f"Row {i}: Value error - {ve}")
                except Exception as e:
                    errors.append(f"Row {i}: {e}")
        
        if success_count:
            messages.success(request, f"{success_count} MTOP records imported successfully!")
        
        if errors:
            messages.warning(request, "Some rows could not be imported:\n" + "\n".join(errors[:10]))
            if len(errors) > 10:
                messages.warning(request, f"... and {len(errors) - 10} more errors.")
    
    except Exception as e:
        messages.error(request, f"Failed to read file: {e}")
    
    return redirect('mtop')


def export_mtop(request):
    """Export MTOP records to CSV or Excel file based on format parameter"""
    export_format = request.GET.get('format', 'csv').lower()
    
    if export_format == 'excel':
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MTOP Records"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Write headers
        headers = [
            'name', 'case_no', 'address', 'no_of_units', 'route_operation',
            'make', 'motor_no', 'chasses_no', 'plate_no', 'date',
            'municipal_treasurer', 'officer_in_charge', 'mayor'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        mtops = Mtop.objects.all().order_by('id')
        for row_num, mtop in enumerate(mtops, 2):
            data = [
                mtop.name,
                mtop.case_no,
                mtop.address,
                mtop.no_of_units,
                mtop.route_operation,
                mtop.make,
                mtop.motor_no,
                mtop.chasses_no,
                mtop.plate_no,
                mtop.date,
                mtop.municipal_treasurer,
                mtop.officer_in_charge,
                mtop.mayor,
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="mtop_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    else:  # Default to CSV
        # Create the HttpResponse object with CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="mtop_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'name', 'case_no', 'address', 'no_of_units', 'route_operation',
            'make', 'motor_no', 'chasses_no', 'plate_no', 'date',
            'municipal_treasurer', 'officer_in_charge', 'mayor'
        ])
        
        # Write data
        mtops = Mtop.objects.all().order_by('id')
        for mtop in mtops:
            writer.writerow([
                mtop.name,
                mtop.case_no,
                mtop.address,
                mtop.no_of_units,
                mtop.route_operation,
                mtop.make,
                mtop.motor_no,
                mtop.chasses_no,
                mtop.plate_no,
                mtop.date.strftime('%Y-%m-%d'),
                mtop.municipal_treasurer,
                mtop.officer_in_charge,
                mtop.mayor,
            ])
        
        return response

def import_franchise(request):
    """Import Franchise records from CSV or Excel file"""
    if request.method != "POST":
        messages.error(request, 'Invalid request method')
        return redirect('franchise')
    
    # Check if file exists in request
    uploaded_file = request.FILES.get('import_file')
    if not uploaded_file:
        messages.error(request, "No file uploaded.")
        return redirect('franchise')
    
    file_name = uploaded_file.name.lower()
    is_excel = file_name.endswith(('.xlsx', '.xls'))
    is_csv = file_name.endswith('.csv')
    
    if not (is_csv or is_excel):
        messages.error(request, "Please upload a CSV or Excel file (.csv, .xlsx, .xls)")
        return redirect('franchise')
    
    errors = []
    success_count = 0
    
    # Expected header columns
    expected_header = ['name', 'denomination', 'plate_no', 'valid_until', 'motor_no',
                      'authorized_no', 'chassis_no', 'authorized_route', 'purpose',
                      'official_receipt_no', 'date', 'amount_paid', 'municipal_treasurer']
    
    def normalize_header(header):
        """Convert header to lowercase, replace spaces/periods with underscores, and fix common variations"""
        normalized = []
        for h in header:
            h = h.lower().strip()
            h = h.replace(' ', '_').replace('.', '')
            h = h.replace('chassis', 'chassis').replace('authorized', 'authorized')
            normalized.append(h)
        return normalized
    
    try:
        if is_excel:
            df = pd.read_excel(uploaded_file)
        else:
            df = pd.read_csv(uploaded_file)
        
        # Normalize DataFrame columns
        df.columns = normalize_header(df.columns)
        
        for idx, row in df.iterrows():
            try:
                # Check for duplicates
                if Franchise.objects.filter(plate_no=row['plate_no']).exists():
                    errors.append(f"Row {idx+2}: Plate No '{row['plate_no']}' already exists.")
                    continue
                
                if Franchise.objects.filter(authorized_no=row['authorized_no']).exists():
                    errors.append(f"Row {idx+2}: Authorized No '{row['authorized_no']}' already exists.")
                    continue
                
                # Convert amount_paid to int (handle decimal strings)
                try:
                    amount_paid = int(float(row.get('amount_paid', 0)))
                except (ValueError, TypeError):
                    amount_paid = 0
                
                # Create franchise record
                Franchise.objects.create(
                    name=row.get('name', ''),
                    denomination=row.get('denomination', ''),
                    plate_no=row.get('plate_no', ''),
                    valid_until=row.get('valid_until', ''),
                    motor_no=row.get('motor_no', ''),
                    authorized_no=row.get('authorized_no', ''),
                    chassis_no=row.get('chassis_no', ''),
                    authorized_route=row.get('authorized_route', ''),
                    purpose=row.get('purpose', ''),
                    official_receipt_no=row.get('official_receipt_no', ''),
                    date=row.get('date', ''),
                    amount_paid=amount_paid,
                    municipal_treasurer=row.get('municipal_treasurer', '')
                )
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {idx+2}: {str(e)}")
        
        if success_count > 0:
            messages.success(request, f"Successfully imported {success_count} franchise record(s).")
        
        if errors:
            for error in errors:
                messages.warning(request, error)
        
        return redirect('franchise')
        
    except Exception as e:
        messages.error(request, f"Error processing file: {str(e)}")
        return redirect('franchise')


def export_franchise(request):
    """Export Franchise records to CSV or Excel file based on format parameter"""
    export_format = request.GET.get('format', 'csv').lower()
    
    if export_format == 'excel':
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Franchise Records"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Write headers
        headers = [
            'name', 'denomination', 'plate_no', 'valid_until', 'motor_no',
            'authorized_no', 'chassis_no', 'authorized_route', 'purpose',
            'official_receipt_no', 'date', 'amount_paid', 'municipal_treasurer'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        franchises = Franchise.objects.all().order_by('id')
        for row_num, franchise in enumerate(franchises, 2):
            data = [
                franchise.name,
                franchise.denomination,
                franchise.plate_no,
                franchise.valid_until,
                franchise.motor_no,
                franchise.authorized_no,
                franchise.chassis_no,
                franchise.authorized_route,
                franchise.purpose,
                franchise.official_receipt_no,
                franchise.date,
                franchise.amount_paid,
                franchise.municipal_treasurer,
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Auto adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
        
        # Save to BytesIO
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="franchise_records.xlsx"'
        wb.save(response)
        return response
    
    else:
        # CSV export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="franchise_records.csv"'
        
        writer = csv.writer(response)
        headers = [
            'name', 'denomination', 'plate_no', 'valid_until', 'motor_no',
            'authorized_no', 'chassis_no', 'authorized_route', 'purpose',
            'official_receipt_no', 'date', 'amount_paid', 'municipal_treasurer'
        ]
        writer.writerow(headers)
        
        franchises = Franchise.objects.all().order_by('id')
        for franchise in franchises:
            writer.writerow([
                franchise.name,
                franchise.denomination,
                franchise.plate_no,
                franchise.valid_until,
                franchise.motor_no,
                franchise.authorized_no,
                franchise.chassis_no,
                franchise.authorized_route,
                franchise.purpose,
                franchise.official_receipt_no,
                franchise.date,
                franchise.amount_paid,
                franchise.municipal_treasurer,
            ])
        
        return response

def mtop_print(request, pk):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    mtop = Mtop.objects.get(pk=pk)
    return render(request, 'myapp/mtop-print.html', {'mtop': mtop})


def mayors_permit_tricycle(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    permits = MayorsPermitTricycle.objects.all()
    return render(request, 'myapp/mayors-permit-tri.html', {'permits': permits})


def mayors_permit_tri_history_data(request, permit_id):
    """API endpoint to fetch status history for a specific tricycle permit"""
    try:
        # Get the permit
        permit = get_object_or_404(MayorsPermitTricycle, id=permit_id)
        
        # Get all history records for this permit
        history_records = MayorsPermitTricycleHistory.objects.filter(
            permit_id=permit_id
        ).order_by('-activated_at')
        
        # Format history data
        history_data = []
        for record in history_records:
            # Determine who changed it
            if record.updated_by_name:
                changed_by = record.updated_by_name
            elif record.updated_by_type:
                changed_by = record.updated_by_type.capitalize()
            else:
                changed_by = 'System'
            
            history_data.append({
                'date': record.activated_at.strftime('%Y-%m-%d %H:%M:%S'),
                'old_status': record.previous_status.capitalize(),
                'new_status': record.new_status.capitalize(),
                'changed_by': changed_by,
                'reason': record.remarks or ''
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'permit_info': {
                    'control_no': permit.control_no,
                    'name': permit.name,
                    'current_status': permit.status
                },
                'history': history_data
            }
        })
        
    except MayorsPermitTricycle.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Permit not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
def mayors_permit_tricycle_datatable(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    
    queryset = MayorsPermitTricycle.objects.select_related('tricycle').all()  # ✅ add select_related
    
    if search_value:
        queryset = queryset.filter(
            Q(control_no__icontains=search_value) |
            Q(name__icontains=search_value) |
            Q(address__icontains=search_value) |
            Q(business_name__icontains=search_value) |
            Q(motorized_operation__icontains=search_value) |
            Q(or_no__icontains=search_value) |
            Q(status__icontains=search_value) |
            Q(issued_at__icontains=search_value) |
            Q(mayor__icontains=search_value) |
            Q(quarter__icontains=search_value) |
            Q(tricycle__body_number__icontains=search_value)  # ✅ ADD THIS
        )
    
    column_filters = {
        0: 'control_no',
        1: 'status',
        2: 'name',
        3: 'address',
        4: 'motorized_operation',
        5: 'business_name',
        6: 'expiry_date',
        7: 'amount_paid',
        8: 'or_no',
        9: 'issue_date',
        10: 'issued_at',
        11: 'mayor',
        12: 'quarter',
        13: 'tricycle__body_number',  # ✅ ADD THIS (new column index 13)
    }
    
    for col_idx, field_name in column_filters.items():
        column_search = request.GET.get(f'columns[{col_idx}][search][value]', '').strip()
        if column_search:
            filter_kwargs = {f'{field_name}__icontains': column_search}
            queryset = queryset.filter(**filter_kwargs)
    
    total_records = MayorsPermitTricycle.objects.count()
    filtered_records = queryset.count()
    
    permits = queryset[start:start + length]
    
    data = []
    for permit in permits:
        body_number = permit.tricycle.body_number if permit.tricycle else ''  # ✅ safe FK access
        
        action_html = f'''
        <div class="btn-group" role="group">
             <button class="btn btn-sm btn-info btn-view" title="View Details" 
                data-id="{permit.id}"
                data-control_no="{permit.control_no}"
                data-status="{permit.status}"
                data-name="{permit.name}"
                data-address="{permit.address}"
                data-motorized_operation="{permit.motorized_operation or ''}"
                data-business_name="{permit.business_name or ''}"
                data-expiry_date="{permit.expiry_date}"
                data-amount_paid="{permit.amount_paid}"
                data-or_no="{permit.or_no}"
                data-issue_date="{permit.issue_date}"
                data-issued_at="{permit.issued_at or ''}"
                data-mayor="{permit.mayor or ''}"
                data-quarter="{permit.get_quarter_display()}"
                data-body_number="{body_number}">
                <i class="fas fa-eye"></i>
            </button>
            <button class="btn btn-sm btn-primary btn-print" title="Print" 
                data-permit-id="{permit.id}"
                data-control-no="{permit.control_no}"
                data-name="{permit.name}"
                data-address="{permit.address}"
                data-motorized="{permit.motorized_operation or ''}"
                data-business="{permit.business_name or ''}"
                data-expiry="{permit.expiry_date}"
                data-quarter="{permit.get_quarter_display()}"
                data-amount="{permit.amount_paid}"
                data-or="{permit.or_no}"
                data-issue="{permit.issue_date}"
                data-issued-at="{permit.issued_at or ''}"
                data-mayor="{permit.mayor or ''}">
                <i class="fas fa-print"></i>
            </button>
            <button type="button" class="btn btn-sm btn-warning btn-update" title="Update"
                data-id="{permit.id}"
                data-control_no="{permit.control_no}"
                data-status="{permit.status}"
                data-name="{permit.name}"
                data-address="{permit.address}"
                data-motorized_operation="{permit.motorized_operation or ''}"
                data-business_name="{permit.business_name or ''}"
                data-expiry_date="{permit.expiry_date}"
                data-amount_paid="{permit.amount_paid}"
                data-or_no="{permit.or_no}"
                data-issue_date="{permit.issue_date}"
                data-issued_at="{permit.issued_at or ''}"
                data-mayor="{permit.mayor or ''}"
                data-quarter="{permit.get_quarter_display()}"
                data-body-number="{body_number}">
                <i class="fas fa-edit"></i>
            </button>
            <button type="button" class="btn btn-sm btn-secondary btn-history" data-id="{permit.id}" title="History">
                <i class="fas fa-history"></i>
            </button>
            <button type="button" class="btn btn-sm btn-danger btn-delete-permit" title="Delete"
                data-id="{permit.id}"
                data-control-no="{permit.control_no}"
                data-name="{permit.name}">
                <i class="fas fa-trash"></i>
            </button>
        </div>
        '''
        
        data.append([
            permit.control_no,                          # 0 - Control No
            permit.status,                              # 1 - Status (hidden)
            permit.name,                                # 2 - Name
            permit.address,                             # 3 - Address
            permit.motorized_operation or '',           # 4 - Motorized Operation
            permit.business_name or '',                 # 5 - Business Name
            permit.expiry_date.strftime('%B ') + str(permit.expiry_date.day) + permit.expiry_date.strftime(', %Y'),  # 6    # 6 - Expiry Date (hidden)
            '{:,.2f}'.format(float(permit.amount_paid)),  # 7 - Amount Paid (hidden)                   # 7 - Amount Paid (hidden)
            permit.or_no,                               # 8 - OR No (hidden)
            permit.issue_date.strftime('%B ') + str(permit.issue_date.day) + permit.issue_date.strftime(', %Y'),    # 9   # 9 - Issue Date (hidden)
            permit.issued_at or '',                     # 10 - Issued At (hidden)
            permit.mayor or '',                         # 11 - Mayor (hidden)
            permit.get_quarter_display(),               # 12 - Quarter (hidden)
            action_html,                                # 13 - Action  ← shifted
            permit.id,                                  # 14 - Hidden ID  ← shifted
            permit.status,                              # 15 - Hidden raw status  ← shifted
            body_number,                                # 16 - Body Number (new) ✅
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })
    
def add_permit_tri(request):
    if request.method == 'POST':
        try:
            control_no = request.POST.get('control_no')
            name = request.POST.get('name')
            address = request.POST.get('address')
            motorized_operation = request.POST.get('motorized_operation')
            business_name = request.POST.get('business_name')
            expiry_date = request.POST.get('expiry_date')
            amount_paid = request.POST.get('amount_paid')
            or_no = request.POST.get('or_no')
            issue_date = request.POST.get('issue_date')
            issued_at = request.POST.get('issued_at')
            mayor = request.POST.get('mayor')
            quarter = request.POST.get('quarter')
            status = request.POST.get('status')
            tricycle_body_number = request.POST.get('tricycle')  # ← NEW

            # Resolve tricycle FK
            tricycle = None
            if tricycle_body_number:
                try:
                    tricycle = Tricycle.objects.get(body_number=tricycle_body_number)
                except Tricycle.DoesNotExist:
                    messages.error(request, f"Tricycle with body number '{tricycle_body_number}' not found.")
                    return redirect('mayors-permit-tricycle')

            permit = MayorsPermitTricycle.objects.create(
                control_no=control_no,
                name=name,
                address=address,
                motorized_operation=motorized_operation,
                business_name=business_name,
                expiry_date=expiry_date,
                amount_paid=int(amount_paid),
                or_no=or_no,
                issue_date=issue_date,
                issued_at=issued_at,
                mayor=mayor,
                quarter=quarter,
                status=status,
                tricycle=tricycle,  # ← NEW
            )

            messages.success(request, f'Permit {control_no} has been successfully added!')
            return redirect('mayors-permit-tricycle')

        except Exception as e:
            messages.error(request, f'Error adding permit: {str(e)}')
            return redirect('mayors-permit-tricycle')

    messages.error(request, 'Invalid request method')
    return redirect('mayors-permit-tricycle')


@require_http_methods(["POST"])
def update_permit_tri(request, permit_id):
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Invalid request method'
        }, status=405)
    
    try:
        permit = get_object_or_404(MayorsPermitTricycle, id=permit_id)
        data = json.loads(request.body)

        # ✅ Save previous values BEFORE updating
        previous_status = permit.status
        previous_expiry_date = permit.expiry_date
        new_status = data.get('status')

        # -------- Update fields --------
        permit.control_no = data.get('control_no')
        permit.name = data.get('name')
        permit.address = data.get('address')
        permit.motorized_operation = data.get('motorized_operation')
        permit.business_name = data.get('business_name')

        if data.get('issue_date'):
            permit.issue_date = datetime.strptime(data.get('issue_date'), '%Y-%m-%d').date()

        new_expiry_date = None
        if data.get('expiry_date'):
            new_expiry_date = datetime.strptime(data.get('expiry_date'), '%Y-%m-%d').date()
            permit.expiry_date = new_expiry_date

        permit.amount_paid = int(data.get('amount_paid', 0))
        permit.or_no = data.get('or_no')
        permit.issued_at = data.get('issued_at')
        permit.mayor = data.get('mayor')
        permit.quarter = data.get('quarter')
        permit.status = new_status

        # Update tricycle FK if body number provided
        tricycle_body_number = data.get('tricycle_body_number', '').strip()
        if tricycle_body_number:
            try:
                permit.tricycle = Tricycle.objects.get(body_number=tricycle_body_number)
            except Tricycle.DoesNotExist:
                return JsonResponse({'success': False, 'error': f"Tricycle with body number '{tricycle_body_number}' not found."}, status=400)
        else:
            permit.tricycle = None

        permit.save()

        # -------- User info --------
        user_type = request.session.get('user_type')
        user_id = request.session.get('superadmin_id') if user_type == 'superadmin' else request.session.get('admin_id')
        user_name = request.session.get('full_name')

        # -------- ✅ MayorsPermitTricycleHistory: status changed --------
        if previous_status != new_status:
            MayorsPermitTricycleHistory.objects.create(
                permit=permit,
                previous_status=previous_status,
                new_status=new_status,
                remarks=f"Status changed from {previous_status} to {new_status}",
                updated_by_type=user_type,
                updated_by_id=user_id,
                updated_by_name=user_name
            )

        # -------- ✅ TricycleHistory: record renewal from permit --------
        if permit.tricycle:
            tricycle = permit.tricycle

            is_renewal = (
                new_status == 'active' and (
                    previous_status != 'active' or
                    (new_expiry_date and new_expiry_date != previous_expiry_date)
                )
            )
            is_status_change = tricycle.status != {
                'active': 'Renewed',
                'inactive': 'Inactive',
                'expired': 'Expired',
            }.get(new_status, tricycle.status)

            if is_renewal:
                TricycleHistory.objects.create(
                    tricycle=tricycle,
                    action='renewed',
                    previous_status=tricycle.status,
                    new_status='Renewed',
                    previous_date_expired=previous_expiry_date,
                    new_date_expired=new_expiry_date or permit.expiry_date,
                    remarks=f"Renewed via Mayor's Permit (Control No: {permit.control_no})",
                    created_by=user_name or 'system'
                )
            elif is_status_change:
                STATUS_MAP = {
                    'active': 'Renewed',
                    'inactive': 'Inactive',
                    'expired': 'Expired',
                }
                new_tricycle_status = STATUS_MAP.get(new_status, tricycle.status)
                TricycleHistory.objects.create(
                    tricycle=tricycle,
                    action='status_changed',
                    previous_status=tricycle.status,
                    new_status=new_tricycle_status,
                    remarks=f"Status updated via Mayor's Permit (Control No: {permit.control_no})",
                    created_by=user_name or 'system'
                )

        messages.success(request, f'Permit for {permit.name} updated successfully!')
        return JsonResponse({'success': True, 'message': 'Permit updated successfully'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_POST
def delete_permit_tri(request):
    try:
        permit_id = request.POST.get('permit_id')

        if not permit_id:
            messages.error(request, 'Permit ID is required')
            return JsonResponse({'success': False, 'error': 'Permit ID is required'})

        permit = MayorsPermitTricycle.objects.get(id=permit_id)

        control_no = permit.control_no
        name = permit.name

        ActivityLog.objects.create(
            action='delete',
            model_type='motorcycle',
            object_id=control_no,
            object_name=name,
            description=f"Mayor's Permit (Tricycle) deleted: {name} (Control No: {control_no})",
            user_type=request.session.get('user_type'),
            user_id=request.session.get('admin_id') or request.session.get('superadmin_id'),
            user_name=request.session.get('full_name'),
        )

        permit.delete()

        messages.success(request, f'Permit "{control_no}" for {name} has been deleted successfully.')
        return JsonResponse({'success': True, 'message': f'Permit "{control_no}" has been deleted successfully.'})

    except MayorsPermitTricycle.DoesNotExist:
        messages.error(request, 'Permit not found.')
        return JsonResponse({'success': False, 'error': 'Permit not found'})
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)})

def get_tricycles(request):
    tricycles = Tricycle.objects.values(
        'body_number',
        'name',
        'address',
        'plate_no',
        'engine_motor_no',
        'chassis_no',
    ).order_by('body_number')
    return JsonResponse({"tricycles": list(tricycles)})

def export_mayors_permit(request):
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mayors Permit"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Excel Header
        headers = [
            'Control No', 'Name', 'Address', 'Motorized Operation',
            'Business Name', 'Expiry Date', 'Amount Paid', 'OR No',
            'Issue Date', 'Issued At', 'Mayor', 'Quarter', 'Status'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Excel Rows
        for row_num, p in enumerate(MayorsPermit.objects.all(), 2):
            ws.cell(row=row_num, column=1, value=p.control_no)
            ws.cell(row=row_num, column=2, value=p.name)
            ws.cell(row=row_num, column=3, value=p.address)
            ws.cell(row=row_num, column=4, value=p.motorized_operation)
            ws.cell(row=row_num, column=5, value=p.business_name)
            ws.cell(row=row_num, column=6, value=str(p.expiry_date) if p.expiry_date else '')
            ws.cell(row=row_num, column=7, value=p.amount_paid)
            ws.cell(row=row_num, column=8, value=p.or_no)
            ws.cell(row=row_num, column=9, value=str(p.issue_date) if p.issue_date else '')
            ws.cell(row=row_num, column=10, value=p.issued_at)
            ws.cell(row=row_num, column=11, value=p.mayor)
            ws.cell(row=row_num, column=12, value=p.quarter)
            ws.cell(row=row_num, column=13, value=p.status)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="mayors_permit.xlsx"'
        wb.save(response)
        return response
    
    else:  # CSV format (default)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="mayors_permit.csv"'
        
        writer = csv.writer(response)
        
        # CSV Header
        writer.writerow([
            'control_no', 'name', 'address', 'motorized_operation',
            'business_name', 'expiry_date', 'amount_paid', 'or_no',
            'issue_date', 'issued_at', 'mayor', 'quarter', 'status'
        ])
        
        # CSV Rows
        for p in MayorsPermit.objects.all():
            writer.writerow([
                p.control_no, p.name, p.address, p.motorized_operation,
                p.business_name, p.expiry_date, p.amount_paid, p.or_no,
                p.issue_date, p.issued_at, p.mayor, p.quarter, p.status
            ])
        
        return response

def export_mayors_permit_tri(request):
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mayors Permit Tricycle"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Excel Header
        headers = [
            'Control No', 'Name', 'Address', 'Motorized Operation',
            'Business Name', 'Expiry Date', 'Amount Paid', 'OR No',
            'Issue Date', 'Issued At', 'Mayor', 'Quarter', 'Status'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Excel Rows
        for row_num, p in enumerate(MayorsPermitTricycle.objects.all(), 2):
            ws.cell(row=row_num, column=1, value=p.control_no)
            ws.cell(row=row_num, column=2, value=p.name)
            ws.cell(row=row_num, column=3, value=p.address)
            ws.cell(row=row_num, column=4, value=p.motorized_operation)
            ws.cell(row=row_num, column=5, value=p.business_name)
            ws.cell(row=row_num, column=6, value=str(p.expiry_date) if p.expiry_date else '')
            ws.cell(row=row_num, column=7, value=p.amount_paid)
            ws.cell(row=row_num, column=8, value=p.or_no)
            ws.cell(row=row_num, column=9, value=str(p.issue_date) if p.issue_date else '')
            ws.cell(row=row_num, column=10, value=p.issued_at)
            ws.cell(row=row_num, column=11, value=p.mayor)
            ws.cell(row=row_num, column=12, value=p.quarter)
            ws.cell(row=row_num, column=13, value=p.status)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="mayors_permit_tricycle.xlsx"'
        wb.save(response)
        return response
    
    else:  # CSV format (default)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="mayors_permit_tricycle.csv"'
        
        writer = csv.writer(response)
        
        # CSV Header
        writer.writerow([
            'control_no', 'name', 'address', 'motorized_operation',
            'business_name', 'expiry_date', 'amount_paid', 'or_no',
            'issue_date', 'issued_at', 'mayor', 'quarter', 'status'
        ])
        
        # CSV Rows
        for p in MayorsPermitTricycle.objects.all():
            writer.writerow([
                p.control_no, p.name, p.address, p.motorized_operation,
                p.business_name, p.expiry_date, p.amount_paid, p.or_no,
                p.issue_date, p.issued_at, p.mayor, p.quarter, p.status
            ])
        
        return response
    
def import_mayors_permit_tri(request):
    if request.method == "POST":
        uploaded_file = request.FILES.get("csv_file")

        if not uploaded_file:
            messages.error(request, "No file uploaded.")
            return redirect("mayors-permit-tricycle")

        file_name = uploaded_file.name.lower()
        is_excel = file_name.endswith((".xlsx", ".xls"))
        is_csv = file_name.endswith(".csv")

        if not (is_csv or is_excel):
            messages.error(request, "Please upload a CSV or Excel file (.csv, .xlsx, .xls)")
            return redirect("mayors-permit-tricycle")

        errors = []
        success_count = 0

        # Expected header columns
        expected_header = [
            "control_no", "name", "address", "motorized_operation",
            "business_name", "expiry_date", "amount_paid", "or_no",
            "issue_date", "issued_at", "mayor", "quarter", "status"
        ]

        def normalize_header(header):
            return [h.strip().lower().replace(" ", "_") for h in header]

        def parse_date_flexible(date_val):
            if isinstance(date_val, datetime):
                return date_val.date()
            if date_val is None or str(date_val).strip() == "":
                return None

            date_str = str(date_val).strip()

            date_formats = [
                "%Y-%m-%d",
                "%m/%d/%Y",
                "%d/%m/%Y",
                "%Y/%m/%d",
                "%m-%d-%Y",
                "%d-%m-%Y",

                "%B %d, %Y",
                "%b %d, %Y",
                "%b %d %Y",
                "%B %d %Y",
                "%d %b %Y",
                "%d %B %Y",

                "%m/%d/%y",
                "%m-%d-%y",

                "%d-%b-%y",
                "%d-%B-%y",
                "%b-%d-%y",
                "%B-%d-%y",
            ]

            for fmt in date_formats:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue

            raise ValueError(
                f"Unable to parse date '{date_val}'. "
                "Supported formats include YYYY-MM-DD, MM/DD/YYYY, DD/MM/YYYY."
            )

        try:
            # ===================== EXCEL =====================
            if is_excel:
                wb = load_workbook(uploaded_file)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))

                if not rows:
                    messages.error(request, "Excel file is empty.")
                    return redirect("mayors-permit-tricycle")

                raw_header = [str(cell).strip() if cell else "" for cell in rows[0]]
                if normalize_header(raw_header) != expected_header:
                    messages.error(
                        request,
                        f"Invalid Excel header. Expected: {', '.join(expected_header)}"
                    )
                    return redirect("mayors-permit-tricycle")

                for i, row in enumerate(rows[1:], start=2):
                    if not row or all(cell is None for cell in row):
                        continue

                    if len(row) < len(expected_header):
                        errors.append(
                            f"Row {i}: Incorrect number of columns ({len(row)})."
                        )
                        continue

                    try:
                        issue_date = parse_date_flexible(row[8])
                        expiry_date = parse_date_flexible(row[5])

                        try:
                            amount_paid = int(float(row[6])) if row[6] else 0
                        except ValueError:
                            raise ValueError(f"Invalid amount_paid '{row[6]}'")

                        MayorsPermitTricycle.objects.update_or_create(
                            control_no=str(row[0]) if row[0] else "",
                            defaults={
                                "name": str(row[1]) if row[1] else "",
                                "address": str(row[2]) if row[2] else "",
                                "motorized_operation": str(row[3]) if row[3] else "",
                                "business_name": str(row[4]) if row[4] else "",
                                "expiry_date": expiry_date,
                                "amount_paid": amount_paid,
                                "or_no": str(row[7]) if row[7] else "",
                                "issue_date": issue_date,
                                "issued_at": str(row[9]) if row[9] else "",
                                "mayor": str(row[10]) if row[10] else "",
                                "quarter": str(row[11]) if row[11] else "",
                                "status": str(row[12]) if row[12] else "active",
                            },
                        )
                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {i}: {e}")

            # ===================== CSV =====================
            else:
                file_data = uploaded_file.read().decode("utf-8-sig").splitlines()
                reader = csv.reader(file_data)
                raw_header = next(reader)

                if normalize_header(raw_header) != expected_header:
                    messages.error(
                        request,
                        f"Invalid CSV header. Expected: {', '.join(expected_header)}"
                    )
                    return redirect("mayors-permit-tricycle")

                for i, row in enumerate(reader, start=2):
                    if len(row) != len(expected_header):
                        errors.append(
                            f"Row {i}: Incorrect number of columns ({len(row)})."
                        )
                        continue

                    try:
                        issue_date = parse_date_flexible(row[8])
                        expiry_date = parse_date_flexible(row[5])

                        try:
                            amount_paid = int(float(row[6])) if row[6] else 0
                        except ValueError:
                            raise ValueError(f"Invalid amount_paid '{row[6]}'")

                        MayorsPermitTricycle.objects.update_or_create(
                            control_no=row[0],
                            defaults={
                                "name": row[1],
                                "address": row[2],
                                "motorized_operation": row[3],
                                "business_name": row[4],
                                "expiry_date": expiry_date,
                                "amount_paid": amount_paid,
                                "or_no": row[7],
                                "issue_date": issue_date,
                                "issued_at": row[9],
                                "mayor": row[10],
                                "quarter": row[11],
                                "status": row[12],
                            },
                        )
                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {i}: {e}")

            if success_count:
                messages.success(
                    request, f"{success_count} permits imported successfully!"
                )

            if errors:
                messages.warning(
                    request,
                    "Some rows could not be imported:\n" + "\n".join(errors[:10])
                )
                if len(errors) > 10:
                    messages.warning(
                        request,
                        f"... and {len(errors) - 10} more errors."
                    )

        except Exception as e:
            messages.error(request, f"Failed to read file: {e}")

        return redirect("mayors-permit-tricycle")

    messages.error(request, "Invalid request method")
    return redirect("mayors-permit-tricycle")

    
def import_mayors_permit(request):
    if request.method == "POST":
        uploaded_file = request.FILES.get("csv_file")

        if not uploaded_file:
            messages.error(request, "No file uploaded.")
            return redirect("mayors-permit")

        file_name = uploaded_file.name.lower()
        is_excel = file_name.endswith(('.xlsx', '.xls'))
        is_csv = file_name.endswith('.csv')

        if not (is_csv or is_excel):
            messages.error(request, "Please upload a CSV or Excel file (.csv, .xlsx, .xls)")
            return redirect("mayors-permit")

        errors = []
        success_count = 0

        # Expected header columns
        expected_header = ['control_no', 'name', 'address', 'motorized_operation', 
                          'business_name', 'expiry_date', 'amount_paid', 'or_no', 
                          'issue_date', 'issued_at', 'mayor', 'quarter', 'status']

        def normalize_header(header):
            """Convert header to lowercase and replace spaces with underscores"""
            return [h.strip().lower().replace(' ', '_') for h in header]
        
        def parse_date_flexible(date_val):
            """Parse date from multiple formats"""
            if isinstance(date_val, datetime):
                return date_val.date()
            elif date_val is None or str(date_val).strip() == '':
                return None
            
            date_str = str(date_val).strip()
            
            # List of common date formats to try
            date_formats = [
                "%Y-%m-%d",      # 2025-01-13
                "%m/%d/%Y",      # 1/13/2025
                "%d/%m/%Y",      # 13/1/2025
                "%Y/%m/%d",      # 2025/1/13
                "%m-%d-%Y",      # 1-13-2025
                "%d-%m-%Y",      # 13-1-2025
            ]
            
            for date_format in date_formats:
                try:
                    return datetime.strptime(date_str, date_format).date()
                except ValueError:
                    continue
            
            # If no format works, raise an error
            raise ValueError(f"Unable to parse date '{date_val}'. Expected formats: YYYY-MM-DD or M/D/YYYY")

        try:
            if is_excel:
                # Handle Excel file
                wb = load_workbook(uploaded_file)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                    messages.error(request, "Excel file is empty.")
                    return redirect("mayors-permit")
                
                # Validate and normalize header
                raw_header = [str(cell).strip() if cell else '' for cell in rows[0]]
                normalized_header = normalize_header(raw_header)
                
                if normalized_header != expected_header:
                    messages.error(request, f"Invalid Excel header. Expected: {', '.join(expected_header)}. Got: {', '.join(normalized_header)}")
                    return redirect('mayors-permit')
                
                # Skip header row
                data_rows = rows[1:]
                
                for i, row in enumerate(data_rows, start=2):
                    if not row or all(cell is None for cell in row):
                        continue
                        
                    if len(row) < 13:
                        errors.append(f"Row {i}: Incorrect number of columns ({len(row)}).")
                        continue

                    try:
                        # Use flexible date parsing for Excel
                        issue_date = parse_date_flexible(row[8])
                        expiry_date = parse_date_flexible(row[5])
                        amount_paid = int(float(row[6])) if row[6] else 0

                        MayorsPermit.objects.update_or_create(
                            control_no=str(row[0]) if row[0] else '',
                            defaults={
                                "name": str(row[1]) if row[1] else '',
                                "address": str(row[2]) if row[2] else '',
                                "motorized_operation": str(row[3]) if row[3] else '',
                                "business_name": str(row[4]) if row[4] else '',
                                "expiry_date": expiry_date,
                                "amount_paid": amount_paid,
                                "or_no": str(row[7]) if row[7] else '',
                                "issue_date": issue_date,
                                "issued_at": str(row[9]) if row[9] else '',
                                "mayor": str(row[10]) if row[10] else '',
                                "quarter": str(row[11]) if row[11] else '',
                                "status": str(row[12]) if row[12] else 'active',
                            },
                        )
                        success_count += 1

                    except ValueError as ve:
                        errors.append(f"Row {i}: Value error - {ve}")
                    except Exception as e:
                        errors.append(f"Row {i}: {e}")

            else:
                # Handle CSV file
                file_data = uploaded_file.read().decode("utf-8").splitlines()
                reader = csv.reader(file_data)
                raw_header = next(reader)
                
                # Validate and normalize header
                normalized_header = normalize_header(raw_header)
                
                if normalized_header != expected_header:
                    messages.error(request, f"Invalid CSV header. Expected: {', '.join(expected_header)}. Got: {', '.join(normalized_header)}")
                    return redirect('mayors-permit')

                for i, row in enumerate(reader, start=2):
                    if len(row) != 13:
                        errors.append(f"Row {i}: Incorrect number of columns ({len(row)}).")
                        continue

                    try:
                        # Use flexible date parsing for CSV
                        issue_date = parse_date_flexible(row[8])
                        expiry_date = parse_date_flexible(row[5])
                        amount_paid = int(float(row[6])) if row[6] else 0

                        MayorsPermit.objects.update_or_create(
                            control_no=row[0],
                            defaults={
                                "name": row[1],
                                "address": row[2],
                                "motorized_operation": row[3],
                                "business_name": row[4],
                                "expiry_date": expiry_date,
                                "amount_paid": amount_paid,
                                "or_no": row[7],
                                "issue_date": issue_date,
                                "issued_at": row[9],
                                "mayor": row[10],
                                "quarter": row[11],
                                "status": row[12],
                            },
                        )
                        success_count += 1

                    except ValueError as ve:
                        errors.append(f"Row {i}: Value error - {ve}")
                    except Exception as e:
                        errors.append(f"Row {i}: {e}")

            if success_count:
                messages.success(request, f"{success_count} permits imported successfully!")

            if errors:
                messages.warning(request, "Some rows could not be imported:\n" + "\n".join(errors[:10]))
                if len(errors) > 10:
                    messages.warning(request, f"... and {len(errors) - 10} more errors.")

        except Exception as e:
            messages.error(request, f"Failed to read file: {e}")

        return redirect("mayors-permit")


def sample_print(request):
    return render(request, 'myapp/sample-print.html')


def mayors_permit_history(request, permit_id):
    permit = get_object_or_404(MayorsPermit, id=permit_id)
    history = permit.histories.order_by('-activated_at')

    return render(request, 'myapp/mayors-permit-history.html', {
        'permit': permit,
        'history': history
    })


def create_report_tri(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    return render(request, 'myapp/create-report-tri.html')

def create_report_tri_datatable(request):
    if not (request.session.get('admin_id') or request.session.get('superadmin_id')):
        return redirect('login')
    """Server-side processing endpoint for Create Report DataTables - SIMPLIFIED VERSION"""
    
    # DataTables parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    
    # Date filter parameters
    filter_start = request.GET.get('filter_start', '')
    filter_end = request.GET.get('filter_end', '')
    
    # Base queryset
    queryset = Tricycle.objects.all()
    
    # ✅ IMPROVED DATE FILTERING - Check BOTH date_registered AND date_expired
    if filter_start and filter_end:
        queryset = queryset.filter(
            Q(date_registered__gte=filter_start, date_registered__lte=filter_end) |
            Q(date_expired__gte=filter_start, date_expired__lte=filter_end)
        )
    elif filter_start:
        queryset = queryset.filter(
            Q(date_registered__gte=filter_start) |
            Q(date_expired__gte=filter_start)
        )
    elif filter_end:
        queryset = queryset.filter(
            Q(date_registered__lte=filter_end) |
            Q(date_expired__lte=filter_end)
        )
    
    # Global search
    if search_value:
        queryset = queryset.filter(
            Q(body_number__icontains=search_value) |
            Q(name__icontains=search_value) |
            Q(address__icontains=search_value) |
            Q(make_kind__icontains=search_value) |
            Q(engine_motor_no__icontains=search_value) |
            Q(chassis_no__icontains=search_value) |
            Q(plate_no__icontains=search_value) |
            Q(status__icontains=search_value) |
            Q(remarks__icontains=search_value) |
            Q(toda__icontains=search_value) | 
            Q(date_registered__icontains=search_value) |
            Q(date_expired__icontains=search_value)
        )
    
    # Column-specific search
    for i in range(12):
        column_search = request.GET.get(f'columns[{i}][search][value]', '')
        if column_search:
            if i == 0:  # Body Number
                queryset = queryset.filter(body_number__icontains=column_search)
            elif i == 1:  # Name
                queryset = queryset.filter(name__icontains=column_search)
            elif i == 2:  # Address
                queryset = queryset.filter(address__icontains=column_search)
            elif i == 3:  # Make/Kind
                queryset = queryset.filter(make_kind__icontains=column_search)
            elif i == 4:  # Engine/Motor No
                queryset = queryset.filter(engine_motor_no__icontains=column_search)
            elif i == 5:  # Chassis No
                queryset = queryset.filter(chassis_no__icontains=column_search)
            elif i == 6:  # Plate No
                queryset = queryset.filter(plate_no__icontains=column_search)
            elif i == 7:  # Date Registered
                queryset = queryset.filter(date_registered__icontains=column_search)
            elif i == 8:  # Date Expired
                queryset = queryset.filter(date_expired__icontains=column_search)
            elif i == 9:  # Status
                queryset = queryset.filter(status__icontains=column_search)
            elif i == 10:  # Remarks
                queryset = queryset.filter(remarks__icontains=column_search)
            elif i == 11:  # TODA
                queryset = queryset.filter(toda__icontains=column_search)
            
    
    # ✅ SIMPLE COUNTS - Based on current status in the FILTERED queryset
    renewed_count = queryset.filter(status='Renewed').count()
    registered_count = queryset.exclude(status='Inactive').count()
    expired_count = queryset.filter(status='Expired').count()
    
    # Total records
    total_records = Tricycle.objects.count()
    filtered_records = queryset.count()
    
    # Pagination
    tricycles = queryset[start:start + length]
    
    # Format data
    data = []
    for tricycle in tricycles:
        
        # Action buttons HTML
        action_html = f'''
        <div class="btn-group" role="group">
         <button class="btn btn-sm btn-warning btn-update" title="Update" 
                data-id="{tricycle.id}"
                data-body-number="{tricycle.body_number}"
                data-name="{tricycle.name}"
                data-address="{tricycle.address}"
                data-make-kind="{tricycle.make_kind}"
                data-engine-motor="{tricycle.engine_motor_no}"
                data-chassis="{tricycle.chassis_no}"
                data-plate-no="{tricycle.plate_no}"
                data-date-registered="{tricycle.date_registered}"
                data-date-expired="{tricycle.date_expired}"
                data-status="{tricycle.status}"
                data-latest-franchise-date="{tricycle.franchises.order_by('-date').first().date.strftime('%b-%d-%Y') if tricycle.franchises.exists() else ''}"
                data-remarks="{tricycle.remarks or ''}"
                data-toda="{tricycle.toda or ''}">
                <i class="fas fa-edit"></i>
            </button>
            <button class="btn btn-sm btn-info btn-view" title="View Details"
                data-id="{tricycle.id}"
                data-body-number="{tricycle.body_number}"
                data-name="{tricycle.name}"
                data-address="{tricycle.address}"
                data-make-kind="{tricycle.make_kind}"
                data-engine-motor="{tricycle.engine_motor_no}"
                data-chassis="{tricycle.chassis_no}"
                data-plate-no="{tricycle.plate_no}"
                data-date-registered="{tricycle.date_registered}"
                data-date-expired="{tricycle.date_expired}"
                data-status="{tricycle.status}"
                data-remarks="{tricycle.remarks or ''}"
                data-toda="{tricycle.toda or ''}">
                <i class="fas fa-eye"></i>
            </button>
            <button class="btn btn-sm btn-danger btn-delete" title="Delete"
                data-id="{tricycle.id}"
                data-body-number="{tricycle.body_number}"
                data-name="{tricycle.name}">
                <i class="fas fa-trash"></i>
            </button>
        </div>
        '''
        
        data.append([
            tricycle.body_number,                                                           # 0
            tricycle.name,                                                                  # 1
            tricycle.address,                                                               # 2
            tricycle.make_kind,                                                             # 3
            tricycle.engine_motor_no,                                                       # 4
            tricycle.chassis_no,                                                            # 5
            tricycle.plate_no,                                                              # 6
            tricycle.date_registered.strftime('%b-') + str(tricycle.date_registered.day) + tricycle.date_registered.strftime('-%Y'),                              # 7
            tricycle.date_expired.strftime('%b-') + str(tricycle.date_expired.day) + tricycle.date_expired.strftime('-%Y') if tricycle.date_expired else '',    # 8
            tricycle.status,                                                                # 9
            tricycle.remarks or '-',                                                        # 10
            tricycle.toda or '',                                                            # 11 - TODA (hidden, searchable)
            action_html,                                                                    # 12
            tricycle.id,                                                                    # 13
        ])
            
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data,
        'statusCounts': {
            'renewed': renewed_count,
            'registered': registered_count,
            'expired': expired_count
        }
    })

def export_create_report_tri(request):
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tricycle Report"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Excel Header
        headers = [
            'Body Number', 'Name', 'Address', 'Make/Kind',
            'Engine/Motor No', 'Chassis No', 'Plate No',
            'Date Registered', 'Date Expired', 'Status', 'Remarks', 'TODA'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Excel Rows
        for row_num, t in enumerate(Tricycle.objects.all(), 2):
            ws.cell(row=row_num, column=1, value=t.body_number)
            ws.cell(row=row_num, column=2, value=t.name)
            ws.cell(row=row_num, column=3, value=t.address)
            ws.cell(row=row_num, column=4, value=t.make_kind)
            ws.cell(row=row_num, column=5, value=t.engine_motor_no)
            ws.cell(row=row_num, column=6, value=t.chassis_no)
            ws.cell(row=row_num, column=7, value=t.plate_no)
            ws.cell(row=row_num, column=8, value=str(t.date_registered) if t.date_registered else '')
            ws.cell(row=row_num, column=9, value=str(t.date_expired) if t.date_expired else '')
            ws.cell(row=row_num, column=10, value=t.status)
            ws.cell(row=row_num, column=11, value=t.remarks or '')
            ws.cell(row=row_num, column=12, value=t.toda or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="tricycle_report.xlsx"'
        wb.save(response)
        return response
    
    else:  # CSV format (default)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="tricycle_report.csv"'
        
        writer = csv.writer(response)
        
        # CSV Header
        writer.writerow([
            'body_number', 'name', 'address', 'make_kind',
            'engine_motor_no', 'chassis_no', 'plate_no',
            'date_registered', 'date_expired', 'status', 'remarks', 'toda'
        ])
        
        # CSV Rows
        for t in Tricycle.objects.all():
            writer.writerow([
                t.body_number, t.name, t.address, t.make_kind,
                t.engine_motor_no, t.chassis_no, t.plate_no,
                t.date_registered, t.date_expired, t.status, t.remarks or '', t.toda or ''  # ✅ added
            ])
                
        return response

def import_create_report_tri(request):
    if request.method == "POST":
        uploaded_file = request.FILES.get("csv_file")

        if not uploaded_file:
            messages.error(request, "No file uploaded.")
            return redirect("create-report-tri")

        file_name = uploaded_file.name.lower()
        is_excel = file_name.endswith(('.xlsx', '.xls'))
        is_csv = file_name.endswith('.csv')

        if not (is_csv or is_excel):
            messages.error(request, "Please upload a CSV or Excel file (.csv, .xlsx, .xls)")
            return redirect("create-report-tri")

        errors = []
        success_count = 0

        expected_header = ['body_number', 'name', 'address', 'make_kind',
                           'engine_motor_no', 'chassis_no', 'plate_no',
                           'date_registered', 'date_expired', 'status', 'remarks', 'toda']

        expected_header_legacy = ['body_number', 'name', 'address', 'make_kind',
                                  'engine_motor_no', 'chassis_no', 'plate_no',
                                  'date_registered', 'date_expired', 'status', 'remarks']

        def normalize_header(header):
            normalized = [h.strip().lower().replace(' ', '_').replace('/', '_') for h in header]
            while normalized and normalized[-1] == '':
                normalized.pop()
            return normalized

        def parse_date(date_val):
            if date_val is None or str(date_val).strip() == '':
                return None
            if isinstance(date_val, datetime):
                return date_val.date()
            import datetime as dt
            if isinstance(date_val, dt.date):
                return date_val
            date_str = str(date_val).strip().replace(' ', '')
            formats = [
                "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y",
                "%b-%d-%y", "%b-%d-%Y", "%B-%d-%y", "%B-%d-%Y",
                "%d-%b-%y", "%d-%b-%Y", "%d-%B-%y", "%d-%B-%Y",
                "%b %d, %Y", "%B %d, %Y", "%m/%d/%y", "%d/%m/%y", "%Y/%m/%d",
            ]
            for fmt in formats:
                try:
                    parsed = datetime.strptime(date_str, fmt)
                    if parsed.year < 100:
                        parsed = parsed.replace(year=parsed.year + 2000)
                    return parsed.date()
                except ValueError:
                    continue
            raise ValueError(f"Cannot parse date: '{date_val}'")

        try:
            if is_excel:
                rows = []
                try:
                    wb = load_workbook(uploaded_file)
                    ws = wb.worksheets[0] if wb.worksheets else None
                    if ws is not None:
                        rows = list(ws.iter_rows(values_only=True))
                    else:
                        raise Exception("No worksheets found")
                except Exception:
                    try:
                        import xlrd
                        uploaded_file.seek(0)
                        wb_xls = xlrd.open_workbook(file_contents=uploaded_file.read())
                        ws_xls = wb_xls.sheet_by_index(0)
                        rows = []
                        for rx in range(ws_xls.nrows):
                            row_vals = []
                            for cx in range(ws_xls.ncols):
                                cell = ws_xls.cell(rx, cx)
                                if cell.ctype == xlrd.XL_CELL_DATE:
                                    date_tuple = xlrd.xldate_as_tuple(cell.value, wb_xls.datemode)
                                    from datetime import date as date_type
                                    row_vals.append(str(date_type(*date_tuple[:3])))
                                else:
                                    row_vals.append(cell.value)
                            rows.append(tuple(row_vals))
                    except Exception as e2:
                        messages.error(request, f"Cannot read Excel file: {str(e2)}. Try saving as .xlsx in Excel via File → Save As.")
                        return redirect("create-report-tri")

                if not rows:
                    messages.error(request, "Excel file is empty.")
                    return redirect("create-report-tri")

                # ✅ Auto-detect header row (skip title row if present)
        # ✅ Auto-detect header row (skip title row if present, search up to row 5)
                normalized_header = []
                header_row_index = None

                for idx in range(min(5, len(rows))):
                    raw_header = [str(cell).strip() if cell else '' for cell in rows[idx]]
                    normalized_header = normalize_header(raw_header)
                    if normalized_header in [expected_header, expected_header_legacy]:
                        header_row_index = idx
                        break

                if header_row_index is None:
                    messages.error(request, f"Could not find a valid header row in the first 5 rows.")
                    return redirect('create-report-tri')

                data_rows = rows[header_row_index + 1:]

                if normalized_header == expected_header:
                    has_toda = True
                elif normalized_header == expected_header_legacy:
                    has_toda = False
                else:
                    messages.error(request, f"Invalid Excel header. Got: {', '.join(normalized_header)}")
                    return redirect('create-report-tri')

                for i, row in enumerate(data_rows, start=header_row_index + 2):
                    if not row or all(cell is None for cell in row):
                        continue

                    row = [str(cell).strip() if cell is not None else '' for cell in row]

                    expected_col_count = len(expected_header) if has_toda else len(expected_header_legacy)
                    row = row[:expected_col_count]  # ✅ Trim extra blank columns

                    # ✅ Pad short rows with empty strings instead of skipping
                    while len(row) < expected_col_count:
                        row.append('')

                    # ✅ Skip fully blank rows
                    if not row[0] or str(row[0]).strip() == '':
                        continue

                    # ✅ Skip rows with missing date_registered
                    if not row[7] or str(row[7]).strip() == '':
                        errors.append(f"Row {i}: Skipped — 'date_registered' is empty.")
                        continue

                    try:
                        date_registered = parse_date(row[7])
                        date_expired = parse_date(row[8])

                        Tricycle.objects.update_or_create(
                            body_number=str(row[0]) if row[0] else '',
                            defaults={
                                "name": str(row[1]) if row[1] else '',
                                "address": str(row[2]) if row[2] else '',
                                "make_kind": str(row[3]) if row[3] else '',
                                "engine_motor_no": str(row[4]) if row[4] else '',
                                "chassis_no": str(row[5]) if row[5] else None,  # ✅ None instead of ''
                                "plate_no": str(row[6]) if row[6] else '',
                                "date_registered": date_registered,
                                "date_expired": date_expired,
                                "status": str(row[9]) if row[9] else 'New',
                                "remarks": str(row[10]) if row[10] else '',
                                "toda": (str(row[11]) if row[11] else None) if has_toda else None,
                            },
                        )
                        success_count += 1

                    except ValueError as ve:
                        errors.append(f"Row {i}: Value error - {ve}. Row data: {list(enumerate(row[:12]))}")
                    except Exception as e:
                        if 'UNIQUE constraint failed: myapp_tricycle.plate_no' in str(e):
                            errors.append(f"Row {i}: Skipped — plate_no '{row[6]}' already exists for a different body number.")
                        else:
                            errors.append(f"Row {i}: {e}")

            else:
                file_data = uploaded_file.read().decode("utf-8-sig").splitlines()
                reader = csv.reader(file_data)
                raw_header = next(reader)

                normalized_header = normalize_header(raw_header)

                if normalized_header == expected_header:
                    has_toda = True
                elif normalized_header == expected_header_legacy:
                    has_toda = False
                else:
                    messages.error(request, f"Invalid CSV header. Got: {', '.join(normalized_header)}")
                    return redirect('create-report-tri')

                expected_col_count = len(expected_header) if has_toda else len(expected_header_legacy)

                for i, row in enumerate(reader, start=2):
                    row = [c.strip() for c in row]
                    row = row[:expected_col_count]  # ✅ Trim extra blank columns

                    if len(row) < expected_col_count:
                        errors.append(f"Row {i}: Incorrect number of columns ({len(row)}). Expected {expected_col_count}.")
                        continue

                    # ✅ Skip fully blank rows
                    if not row[0]:
                        continue

                    # ✅ Skip rows with missing date_registered
                    if not row[7]:
                        errors.append(f"Row {i}: Skipped — 'date_registered' is empty.")
                        continue

                    try:
                        date_registered = parse_date(row[7])
                        date_expired = parse_date(row[8])

                        Tricycle.objects.update_or_create(
                            body_number=row[0],
                            defaults={
                                "name": row[1],
                                "address": row[2],
                                "make_kind": row[3],
                                "engine_motor_no": row[4],
                                "chassis_no": row[5] if row[5] else None,  # ✅ None instead of ''
                                "plate_no": row[6],
                                "date_registered": date_registered,
                                "date_expired": date_expired,
                                "status": row[9],
                                "remarks": row[10],
                                "toda": (row[11] if row[11] else None) if has_toda else None,
                            },
                        )
                        success_count += 1

                    except ValueError as ve:
                        errors.append(f"Row {i}: Value error - {ve}")
                    except Exception as e:
                        if 'UNIQUE constraint failed: myapp_tricycle.plate_no' in str(e):
                            errors.append(f"Row {i}: Skipped — plate_no '{row[6]}' already exists for a different body number.")
                        else:
                            errors.append(f"Row {i}: {e}")

            if success_count:
                messages.success(request, f"{success_count} tricycle records imported successfully!")

            if errors:
                messages.warning(request, "Some rows could not be imported:\n" + "\n".join(errors[:10]))
                if len(errors) > 10:
                    messages.warning(request, f"... and {len(errors) - 10} more errors.")

        except Exception as e:
            messages.error(request, f"Failed to read file: {e}")

        return redirect("create-report-tri")

    messages.error(request, 'Invalid request method')
    return redirect('create-report-tri')


@require_POST
def add_tricycle(request):
    try:
        body_number = request.POST.get('body_number')
        name = request.POST.get('name')
        address = request.POST.get('address')
        make_kind = request.POST.get('make_kind')
        engine_motor_no = request.POST.get('engine_motor_no')
        chassis_no = request.POST.get('chassis_no')
        plate_no = request.POST.get('plate_no')
        date_registered = request.POST.get('date_registered')
        date_expired = request.POST.get('date_expired') or None
        status = request.POST.get('status')
        remarks = request.POST.get('remarks', '')
        toda = request.POST.get('toda', '')

        # Simple validation
        if not all([body_number, name, address, make_kind, plate_no, date_registered,status]):
            messages.error(request, 'Please fill all required fields.')
            return JsonResponse({'success': False, 'error': 'Please fill all required fields.'})

        tricycle = Tricycle.objects.create(
            body_number=body_number,
            name=name,
            address=address,
            make_kind=make_kind,
            engine_motor_no=engine_motor_no,
            chassis_no=chassis_no,
            plate_no=plate_no,
            date_registered=date_registered,
            date_expired=date_expired,
            status=status,
            remarks=remarks,
            toda=toda or None,
        )

        # ✅ Manually log to ActivityLog
        ActivityLog.objects.create(
            action='create',
            model_type='tricycle',
            object_id=tricycle.body_number,
            object_name=tricycle.name,
            description=f'New tricycle registered for {tricycle.name} (Body # {tricycle.body_number})',
            user_type=request.session.get('user_type'),
            user_id=request.session.get('admin_id') or request.session.get('superadmin_id'),
            user_name=request.session.get('full_name'),
        )
        
        # ✅ Record creation history
        TricycleHistory.objects.create(
            tricycle=tricycle,
            action='created',
            new_status=status,
            new_date_expired=date_expired,
            remarks=f'Tricycle created with status: {status}',
            created_by=request.user.username if request.user.is_authenticated else 'system'
        )
        
        messages.success(request, f"Tricycle {tricycle.body_number} added successfully!")
        return JsonResponse({'success': True})

    except IntegrityError as e:
        messages.error(request, 'Body Number or Plate Number already exists')
        return JsonResponse({'success': False, 'error': 'Body Number or Plate Number already exists'})
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)})
    
@require_http_methods(["POST"])
def update_tricycle(request):
    import datetime
    """Update an existing tricycle record"""
    try:
        tricycle_id = request.POST.get('tricycle_id')
        
        if not tricycle_id:
            messages.error(request, 'Tricycle ID is required')
            return JsonResponse({'success': False, 'error': 'Tricycle ID is required'})
        
        try:
            tricycle = Tricycle.objects.get(id=tricycle_id)
        except Tricycle.DoesNotExist:
            messages.error(request, 'Tricycle not found')
            return JsonResponse({'success': False, 'error': 'Tricycle not found'})
        
        # Store old values BEFORE updating
        old_status = tricycle.status
        old_date_expired = tricycle.date_expired
        
        # Get form data
        body_number = request.POST.get('body_number')
        name = request.POST.get('name')
        address = request.POST.get('address')
        make_kind = request.POST.get('make_kind')
        engine_motor_no = request.POST.get('engine_motor_no', '')
        chassis_no = request.POST.get('chassis_no', '')
        plate_no = request.POST.get('plate_no')
        date_registered = request.POST.get('date_registered')
        date_expired = request.POST.get('date_expired')
        status = request.POST.get('status')
        remarks = request.POST.get('remarks', '')
        toda = request.POST.get('toda', '')
        
        # Validate required fields
        if not all([body_number, name, address, make_kind, plate_no, date_registered, date_expired, status]):
            messages.error(request, 'All required fields must be filled')
            return JsonResponse({'success': False, 'error': 'All required fields must be filled'})
        
        # Check for unique constraints
        if Tricycle.objects.filter(body_number=body_number).exclude(id=tricycle_id).exists():
            messages.error(request, f'Body Number "{body_number}" already exists')
            return JsonResponse({'success': False, 'error': 'Body Number already exists'})
        
        if Tricycle.objects.filter(plate_no=plate_no).exclude(id=tricycle_id).exists():
            messages.error(request, f'Plate Number "{plate_no}" already exists')
            return JsonResponse({'success': False, 'error': 'Plate Number already exists'})
        
        # Update the tricycle object
        tricycle.body_number = body_number
        tricycle.name = name
        tricycle.address = address
        tricycle.make_kind = make_kind
        tricycle.engine_motor_no = engine_motor_no
        tricycle.chassis_no = chassis_no
        tricycle.plate_no = plate_no
        tricycle.date_registered = datetime.date.fromisoformat(date_registered)
        tricycle.date_expired = datetime.date.fromisoformat(date_expired)
        tricycle.status = status
        tricycle.remarks = remarks
        tricycle.toda = toda or None

        # Attach current user for the signal
        tricycle._current_user = {
            'type': request.session.get('user_type'),
            'id': request.session.get('admin_id') or request.session.get('superadmin_id'),
            'name': request.session.get('full_name'),
        }

        tricycle.save()

        # ✅ Compute what changed FIRST — right after save
        status_changed = old_status != status
        date_expired_changed = str(old_date_expired) != date_expired


        # ✅ Record Tricycle history if anything changed
        if status_changed or date_expired_changed:
            if status == 'Renewed' and old_status != 'Renewed':
                action = 'renewed'
                history_remarks = f'Tricycle renewed from {old_status} to {status}'
            elif status == 'Expired' and old_status != 'Expired':
                action = 'expired'
                history_remarks = f'Tricycle expired from {old_status}'
            elif status_changed:
                action = 'status_changed'
                history_remarks = f'Status changed from {old_status} to {status}'
            else:
                action = 'updated'
                history_remarks = f'Expiry date updated from {old_date_expired} to {date_expired}'
            
            TricycleHistory.objects.create(
                tricycle=tricycle,
                action=action,
                previous_status=old_status if status_changed else None,
                new_status=status if status_changed else None,
                previous_date_expired=old_date_expired if date_expired_changed else None,
                new_date_expired=date_expired if date_expired_changed else None,
                remarks=history_remarks,
                created_by=request.user.username if request.user.is_authenticated else 'system'
            )
        
        messages.success(request, f'Tricycle "{body_number}" updated successfully!')
        return JsonResponse({'success': True, 'message': 'Tricycle updated successfully'})
        
    except IntegrityError as e:
        messages.error(request, f'Database error: {str(e)}')
        return JsonResponse({'success': False, 'error': f'Database error: {str(e)}'})
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return JsonResponse({'success': False, 'error': f'An error occurred: {str(e)}'})


@require_POST
def delete_tricycle(request):
    try:
        tricycle_id = request.POST.get('tricycle_id')
        
        if not tricycle_id:
            messages.error(request, 'Tricycle ID is required')
            return JsonResponse({'success': False, 'error': 'Tricycle ID is required'})
        
        tricycle = Tricycle.objects.get(id=tricycle_id)
        
        body_number = tricycle.body_number
        name = tricycle.name

        ActivityLog.objects.create(
            action='delete',
            model_type='tricycle',
            object_id=body_number,
            object_name=name,
            description=f'Tricycle deleted: {name} (Body # {body_number})',
            user_type=request.session.get('user_type'),
            user_id=request.session.get('admin_id') or request.session.get('superadmin_id'),
            user_name=request.session.get('full_name'),
        )
        
        tricycle.delete()
        
        messages.success(request, f'Tricycle "{body_number}" has been deleted successfully.')
        return JsonResponse({'success': True, 'message': f'Tricycle "{body_number}" has been deleted successfully.'})
    
    except Tricycle.DoesNotExist:
        messages.error(request, 'Tricycle not found.')
        return JsonResponse({'success': False, 'error': 'Tricycle not found'})
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)})

    
def debug_storage(request):
    from django.conf import settings
    info = {
        'DEFAULT_FILE_STORAGE': getattr(settings, 'DEFAULT_FILE_STORAGE', 'NOT SET'),
        'CLOUDINARY_CLOUD_NAME': 'SET' if settings.CLOUDINARY_STORAGE.get('CLOUD_NAME') else 'NOT SET',
        'CLOUDINARY_API_KEY': 'SET' if settings.CLOUDINARY_STORAGE.get('API_KEY') else 'NOT SET',
        'CLOUDINARY_API_SECRET': 'SET' if settings.CLOUDINARY_STORAGE.get('API_SECRET') else 'NOT SET',
        'INSTALLED_APPS': list(settings.INSTALLED_APPS),
    }
    from django.http import JsonResponse
    return JsonResponse(info)